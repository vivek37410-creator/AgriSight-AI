import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "leaf_doctor_training.xlsx")

HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0'),
)


def style_header(ws, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def add_sheet_data(ws, headers, rows):
    ws.append(headers)
    style_header(ws, len(headers))
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = ALT_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = Alignment(vertical='top', wrap_text=True)
    auto_width(ws)


wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: Leaf Disease Training Data
ws_leaf = wb.create_sheet("Leaf Diseases")
leaf_headers = [
    "Crop", "Condition", "Symptoms", "Causes", "Risk Factors",
    "Prevention", "Recommended Action", "Severity", "Model Note"
]
leaf_data = [
    ["Rice", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Rice", "blast",
     "Diamond-shaped lesions on leaves, neck blast causes white or grayish neck.",
     "Fungus Magnaporthe oryzae, favored by cool wet nights and high nitrogen.",
     "High nitrogen fertilization, cool nights with dew, susceptible varieties.",
     "Use resistant varieties, balanced nitrogen, avoid late planting.",
     "Apply recommended fungicide. Avoid excess nitrogen. Ensure good field drainage.",
     "HIGH",
     "Common in high humidity and moderate temperatures."],
    ["Rice", "brown_spot",
     "Oval brown spots on leaves, may merge, causing leaf drying.",
     "Fungus Cochliobolus miyabeanus, thrives in humid conditions.",
     "High humidity, poor soil fertility, potassium deficiency.",
     "Use resistant varieties, balanced fertilization, seed treatment.",
     "Apply recommended fungicide. Improve soil fertility. Use healthy seed.",
     "MODERATE",
     "Often associated with nutrient-deficient soils."],
    ["Rice", "sheath_blight",
     "Elliptical lesions on leaf sheaths, lodging in severe cases.",
     "Fungus Rhizoctonia solani, spreads in high humidity.",
     "Dense planting, high humidity, excessive nitrogen.",
     "Use resistant varieties, proper spacing, balanced nitrogen.",
     "Apply recommended fungicide. Improve air circulation. Avoid excess nitrogen.",
     "MODERATE",
     "More common in densely planted fields."],
    ["Wheat", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Wheat", "rust",
     "Orange or brown pustules on leaves and stems, premature ripening.",
     "Fungal pathogens Puccinia spp., wind-dispersed spores.",
     "Cool moist weather, susceptible varieties, dense sowing.",
     "Grow resistant varieties, timely sowing, monitor regularly.",
     "Apply recommended fungicide at early signs. Remove volunteer wheat. Use resistant varieties.",
     "HIGH",
     "Spreads rapidly in cool, moist conditions."],
    ["Wheat", "powdery_mildew",
     "White powdery fungal growth on leaves and stems.",
     "Fungus Blumeria graminis, thrives in high humidity and moderate temperatures.",
     "High humidity, dense canopy, susceptible varieties.",
     "Use resistant varieties, proper spacing, avoid excess nitrogen.",
     "Apply recommended fungicide. Improve air circulation. Remove infected leaves.",
     "MODERATE",
     "Common in dense canopies with poor air circulation."],
    ["Maize", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Maize", "northern_leaf_blight",
     "Long cigar-shaped grayish lesions on leaves, leaf drying.",
     "Fungus Exserohilum turcicum, spread by wind and rain.",
     "High humidity, moderate temperatures, susceptible hybrids.",
     "Grow resistant hybrids, crop rotation, residue management.",
     "Apply recommended fungicide. Practice crop rotation. Remove infected debris.",
     "MODERATE",
     "Common in temperate maize growing regions."],
    ["Maize", "common_rust",
     "Small reddish-brown pustules on leaf surfaces, premature leaf drying.",
     "Fungus Puccinia sorghi, wind-borne spores.",
     "High humidity, moderate temperatures, susceptible varieties.",
     "Use resistant varieties, early planting, scout regularly.",
     "Apply recommended fungicide at early detection. Monitor fields regularly.",
     "MODERATE",
     "Spreads rapidly in warm, humid conditions."],
    ["Tomato", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Tomato", "early_blight",
     "Dark brown spots with concentric rings, yellowing leaves, defoliation.",
     "Fungus Alternaria solani, survives in plant debris.",
     "High humidity, warm temperatures, poor air circulation.",
     "Use resistant varieties, crop rotation, proper spacing.",
     "Apply recommended fungicide. Remove infected leaves. Improve air circulation.",
     "HIGH",
     "Common in warm, humid conditions."],
    ["Tomato", "late_blight",
     "Water-soaked spots on leaves, white fungal growth on undersides, fruit rot.",
     "Oomycete Phytophthora infestans, thrives in cool moist conditions.",
     "Cool wet weather, poor air circulation, overhead irrigation.",
     "Use resistant varieties, avoid overhead irrigation, good air circulation.",
     "Apply recommended fungicide immediately. Remove infected plants. Avoid wetting foliage.",
     "HIGH",
     "Rapidly spreading in cool, wet conditions."],
    ["Cotton", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Cotton", "bacterial_blight",
     "Angular brown spots on leaves, black veins, leaf dropping.",
     "Bacteria Xanthomonas citri pv. malvacearum, spreads through water.",
     "High humidity, warm temperatures, infected seed.",
     "Use resistant varieties, certified seed, crop rotation.",
     "Apply recommended bactericide. Remove infected plants. Use certified seed.",
     "HIGH",
     "Spread through rain splash and contaminated tools."],
    ["Cotton", "alternaria_leaf_spot",
     "Brown circular spots with concentric rings, leaf drying.",
     "Fungus Alternaria alternata, favored by warm humid weather.",
     "High humidity, warm temperatures, plant stress.",
     "Use resistant varieties, balanced fertilization, proper spacing.",
     "Apply recommended fungicide. Remove infected debris. Improve plant nutrition.",
     "MODERATE",
     "Often appears under plant stress conditions."],
    ["Soybean", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Soybean", "frogeye_leaf_spot",
     "Small brown spots with purple margins on upper leaf surface.",
     "Fungus Cercospora sojina, survives in crop residue.",
     "High humidity, warm temperatures, continuous soybean.",
     "Use resistant varieties, crop rotation, residue management.",
     "Apply recommended fungicide. Practice crop rotation. Use resistant varieties.",
     "MODERATE",
     "Common in continuous soybean systems."],
    ["Soybean", "soybean_rust",
     "Small reddish-brown pustules on leaf undersides, premature leaf drop.",
     "Fungus Phakopsora pachyrhizi, wind-borne spores.",
     "Warm humid conditions, susceptible varieties, continuous soybean.",
     "Use resistant varieties, early planting, scout regularly.",
     "Apply recommended fungicide at early detection. Monitor fields regularly after flowering.",
     "HIGH",
     "Serious threat in warm, humid conditions."],
    ["Groundnut", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Groundnut", "late_leaf_spot",
     "Brown to dark spots on lower leaves, leaf defoliation.",
     "Fungus Cercosporidium personatum, favored by humid conditions.",
     "High humidity, dense canopy, susceptible varieties.",
     "Use tolerant varieties, proper spacing, fungicide at peg initiation.",
     "Apply recommended fungicide. Improve air circulation through proper spacing.",
     "MODERATE",
     "Major yield-limiting disease in groundnut."],
    ["Groundnut", "rust",
     "Orange to brown pustules on leaf undersides, leaf drying.",
     "Fungus Puccinia arachidis, spreads in warm humid weather.",
     "High humidity, warm temperatures, susceptible varieties.",
     "Use resistant varieties, timely sowing, monitor regularly.",
     "Apply recommended fungicide. Use resistant varieties. Practice crop rotation.",
     "MODERATE",
     "Common in warm, humid groundnut fields."],
    ["Chickpea", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Chickpea", "wilt",
     "Yellowing and wilting of plants from base, brown vascular bundles.",
     "Fungus Fusarium oxysporum, soil-borne pathogen.",
     "High soil temperature, continuous chickpea, susceptible varieties.",
     "Use resistant varieties, seed treatment, crop rotation.",
     "Apply recommended fungicide seed treatment. Practice 3-4 year crop rotation. Use tolerant varieties.",
     "MODERATE",
     "Soil-borne disease, difficult to manage once established."],
    ["Chickpea", "blight",
     "Water-soaked spots on leaves and pods, brown lesions, plant death.",
     "Fungus Ascochyta rabiei, spreads through rain splash.",
     "High humidity, cool temperatures, dense canopy.",
     "Use resistant varieties, seed treatment, proper spacing.",
     "Apply recommended fungicide. Use disease-free seed. Practice crop rotation.",
     "HIGH",
     "Can cause complete crop failure under favorable conditions."],
    ["Pigeon Pea", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Pigeon Pea", "yellow_mosaic",
     "Yellow mosaic patches on leaves, stunted growth, reduced yield.",
     "Virus transmitted by whiteflies.",
     "High whitefly population, warm humid weather, close planting.",
     "Use virus-free seed, control whiteflies, remove infected plants.",
     "Apply recommended insecticide for whitefly control. Remove and destroy infected plants. Use tolerant varieties.",
     "MODERATE",
     "Virus disease spread by whitefly vectors."],
    ["Pigeon Pea", "cercospora_leaf_spot",
     "Circular brown spots with gray centers, leaf yellowing.",
     "Fungus Cercospora cajani, favored by high humidity.",
     "High humidity, warm temperatures, susceptible varieties.",
     "Use resistant varieties, proper spacing, crop rotation.",
     "Apply recommended fungicide. Remove infected plant debris. Use tolerant varieties.",
     "MODERATE",
     "Common in warm, humid pigeon pea fields."],
    ["Sugarcane", "healthy",
     "Uniform green color, no spots or lesions, normal growth.",
     "No disease present.",
     "N/A",
     "Maintain proper spacing, balanced fertilization, and good drainage.",
     "Continue regular monitoring and good agronomic practices.",
     "LOW",
     "Healthy leaf - no action required."],
    ["Sugarcane", "red_rot",
     "Red discoloration in stalk, external white patches, sour smell.",
     "Fungus Colletotrichum falcatum, enters through wounds.",
     "Warm humid weather, poor drainage, ratoon cropping.",
     "Use resistant varieties, hot water treatment of setts, crop rotation.",
     "Remove and destroy infected plants. Treat setts before planting. Avoid ratoon cropping.",
     "HIGH",
     "Serious disease causing significant yield loss."],
    ["Sugarcane", "smut",
     "Whip-like fungal growth from growing point, poor cane development.",
     "Fungus Sporisorium scitamineum, spreads through wind-borne spores.",
     "High humidity, susceptible varieties, infected planting material.",
     "Use resistant varieties, hot water treatment, rogue infected plants.",
     "Remove and destroy infected plants. Use resistant varieties. Treat setts before planting.",
     "HIGH",
     "Systemic disease requiring strict quarantine measures."],
]
add_sheet_data(ws_leaf, leaf_headers, leaf_data)

# Sheet 2: Symptom-Based Lookup
ws_symptoms = wb.create_sheet("Symptom Lookup")
symptom_headers = ["Symptom Keywords", "Likely Condition", "Crop", "Recommended Action"]
symptom_data = [
    ["yellow leaves, mosaic, stunted", "Yellow Mosaic Disease", "Pigeon Pea", "Control whiteflies, use virus-free seed, remove infected plants."],
    ["brown spots, concentric rings, defoliation", "Early Blight", "Tomato", "Apply fungicide, remove infected leaves, improve air circulation."],
    ["water-soaked spots, white mold, fruit rot", "Late Blight", "Tomato", "Apply fungicide immediately, remove infected plants, avoid overhead irrigation."],
    ["orange pustules, brown spots, premature ripening", "Rust", "Wheat", "Apply fungicide, use resistant varieties, remove volunteer wheat."],
    ["cigar-shaped lesions, leaf drying", "Northern Corn Leaf Blight", "Maize", "Apply fungicide, practice crop rotation, remove infected debris."],
    ["red pustules, leaf drop", "Soybean Rust", "Soybean", "Apply fungicide at early detection, use resistant varieties, monitor regularly."],
    ["brown spots, leaf drop, defoliation", "Late Leaf Spot", "Groundnut", "Apply fungicide, improve air circulation, use tolerant varieties."],
    ["wilting, yellowing, brown vascular", "Wilt", "Chickpea", "Apply fungicide seed treatment, practice crop rotation, use tolerant varieties."],
    ["water-soaked spots, brown lesions, death", "Blight", "Chickpea", "Apply fungicide, use disease-free seed, practice crop rotation."],
    ["red discoloration, white patches, sour smell", "Red Rot", "Sugarcane", "Remove infected plants, treat setts, avoid ratoon cropping."],
    ["whip-like growth, poor development", "Smut", "Sugarcane", "Remove infected plants, use resistant varieties, treat setts."],
    ["angular spots, black veins, dropping", "Bacterial Blight", "Cotton", "Apply bactericide, remove infected plants, use certified seed."],
    ["diamond-shaped lesions, neck blast", "Blast", "Rice", "Apply fungicide, avoid excess nitrogen, ensure good drainage."],
    ["oval brown spots, leaf drying", "Brown Spot", "Rice", "Apply fungicide, improve soil fertility, use healthy seed."],
    ["elliptical lesions, lodging", "Sheath Blight", "Rice", "Apply fungicide, improve air circulation, avoid excess nitrogen."],
]
add_sheet_data(ws_symptoms, symptom_headers, symptom_data)

# Sheet 3: Weather-Based Disease Risk
ws_weather = wb.create_sheet("Weather Disease Risk")
weather_headers = ["Crop", "Temperature Range", "Humidity Range", "Rainfall Condition", "Disease Risk", "Precaution"]
weather_data = [
    ["Rice", "20-30°C", ">80%", "Heavy rain / standing water", "HIGH", "Apply preventive fungicide. Ensure proper drainage. Monitor for blast and sheath blight."],
    ["Rice", "25-35°C", ">70%", "Moderate rain", "MODERATE", "Monitor for brown spot and leaf blast. Apply fungicide if needed."],
    ["Wheat", "15-25°C", ">70%", "High humidity / dew", "HIGH", "Apply preventive fungicide. Monitor for rust and powdery mildew."],
    ["Wheat", "10-20°C", ">60%", "Light rain / dew", "MODERATE", "Monitor for rust. Apply fungicide at first signs."],
    ["Maize", "20-30°C", ">75%", "Moderate to heavy rain", "HIGH", "Apply preventive fungicide. Monitor for northern leaf blight."],
    ["Maize", "25-35°C", ">70%", "High humidity", "MODERATE", "Monitor for common rust. Apply fungicide if needed."],
    ["Tomato", "20-28°C", ">80%", "Heavy rain / overhead irrigation", "HIGH", "Apply preventive fungicide. Remove infected plants. Avoid overhead irrigation."],
    ["Tomato", "22-30°C", ">70%", "Moderate humidity", "MODERATE", "Monitor for early blight. Apply fungicide at first signs."],
    ["Cotton", "25-32°C", ">70%", "High humidity / rain", "HIGH", "Apply preventive fungicide. Monitor for bacterial blight."],
    ["Cotton", "28-35°C", ">60%", "Moderate humidity", "MODERATE", "Monitor for Alternaria leaf spot. Apply fungicide if needed."],
    ["Soybean", "22-30°C", ">75%", "Heavy rain", "HIGH", "Apply preventive fungicide. Monitor for soybean rust."],
    ["Soybean", "25-32°C", ">65%", "Moderate humidity", "MODERATE", "Monitor for frogeye leaf spot. Apply fungicide if needed."],
    ["Groundnut", "25-30°C", ">70%", "High humidity", "HIGH", "Apply preventive fungicide. Monitor for late leaf spot."],
    ["Groundnut", "28-35°C", ">60%", "Moderate humidity", "MODERATE", "Monitor for rust. Apply fungicide if needed."],
    ["Chickpea", "15-25°C", ">60%", "Cool humid weather", "HIGH", "Apply preventive fungicide. Monitor for blight and wilt."],
    ["Chickpea", "18-28°C", ">50%", "Moderate humidity", "MODERATE", "Monitor for blight. Apply fungicide at first signs."],
    ["Pigeon Pea", "25-35°C", ">60%", "Warm humid weather", "MODERATE", "Control whiteflies. Monitor for yellow mosaic and leaf spot."],
    ["Sugarcane", "25-35°C", ">70%", "Heavy rain / waterlogging", "HIGH", "Improve drainage. Monitor for red rot. Remove infected plants."],
    ["Sugarcane", "28-38°C", ">60%", "Moderate humidity", "MODERATE", "Monitor for smut. Use resistant varieties."],
]
add_sheet_data(ws_weather, weather_headers, weather_data)

# Sheet 4: Nutrient Deficiency Symptoms
ws_nutrient = wb.create_sheet("Nutrient Deficiency")
nutrient_headers = ["Crop", "Deficiency", "Symptoms", "Affected Plant Part", "Recommended Action"]
nutrient_data = [
    ["General", "Nitrogen Deficiency", "Pale green to yellow leaves, stunted growth, older leaves affected first.", "Older leaves", "Apply nitrogen fertilizer. Use legume rotation."],
    ["General", "Phosphorus Deficiency", "Purple or bronze leaves, poor root development, delayed maturity.", "Older leaves", "Apply phosphorus fertilizer. Use mycorrhizal inoculants."],
    ["General", "Potassium Deficiency", "Leaf edge scorching, weak stems, reduced fruit quality.", "Leaf edges", "Apply potassium fertilizer. Use organic matter mulches."],
    ["General", "Magnesium Deficiency", "Yellowing between leaf veins, leaf curling, premature leaf drop.", "Older leaves", "Apply magnesium sulfate. Use dolomitic lime."],
    ["General", "Calcium Deficiency", "Distorted new leaves, tip burn, poor fruit development.", "New leaves", "Apply calcium fertilizer. Maintain consistent soil moisture."],
    ["General", "Iron Deficiency", "Yellowing between leaf veins, young leaves affected first.", "Young leaves", "Apply iron chelate. Improve soil drainage."],
    ["Rice", "Nitrogen Deficiency", "Pale green leaves, reduced tillering, early senescence.", "Older leaves", "Apply urea in split doses. Use leaf color chart for timing."],
    ["Wheat", "Nitrogen Deficiency", "Pale green leaves, reduced tillering, low protein grain.", "Older leaves", "Apply nitrogen at tillering and jointing stages."],
    ["Maize", "Nitrogen Deficiency", "V-shaped yellowing on leaf tips, stunted growth.", "Older leaves", "Apply nitrogen at planting and side-dress at knee-high."],
    ["Tomato", "Calcium Deficiency", "Blossom end rot, fruit rot, distorted leaves.", "Fruits and new leaves", "Apply calcium nitrate. Maintain consistent watering."],
    ["Cotton", "Potassium Deficiency", "Leaf edge scorching, reduced boll size, leaf dropping.", "Leaf edges", "Apply muriate of potash. Use potassium-rich fertilizer."],
    ["Soybean", "Molybdenum Deficiency", "Yellowish leaves, poor nodulation, reduced yield.", "Whole plant", "Apply sodium molybdate seed treatment."],
]
add_sheet_data(ws_nutrient, nutrient_headers, nutrient_data)

wb.save(EXCEL_PATH)
print(f"Leaf doctor training Excel created at: {EXCEL_PATH}")
print(f"Sheets: {wb.sheetnames}")

# Sheet 5: Leaf Analysis History
ws_history = wb.create_sheet("Leaf Analysis History")
history_headers = [
    "Timestamp", "User ID", "Farm ID", "Farm Name", "Crop", "Condition",
    "Severity", "Health Status", "Risk Score", "Risk Level",
    "Recommendation", "Recommendation Source", "Latitude", "Longitude",
    "Temperature", "Humidity", "Rainfall", "NDVI", "Model Version"
]
history_data = []
add_sheet_data(ws_history, history_headers, history_data)

wb.save(EXCEL_PATH)
print(f"Sheets after adding history: {wb.sheetnames}")
