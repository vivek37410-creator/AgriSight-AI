"""
Generate agrisight_knowledge.xlsx with ~1500 agricultural Q&A entries.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

random.seed(42)

OUTPUT = "C:/Users/DELL/file/OneDrive/Desktop/PROJECT/AgriPlus AI/backend/app/knowledge/agrisight_knowledge.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FAQ"

headers = ["ID", "Category", "Crop", "Topic", "Question", "Keywords", "Answer", "Recommendation", "Severity", "Language"]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E28", end_color="1F4E28", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

crops = [
    "Pigeon Pea", "Rice", "Wheat", "Maize", "Cotton", "Tomato", "Potato", "Onion",
    "Soybean", "Groundnut", "Chickpea", "Sugarcane", "Millet", "Sorghum", "Banana",
    "Mango", "Grapes", "Chilli", "Brinjal", "Okra", "Cabbage", "Cauliflower",
    "Cucumber", "Peas", "Apple", "Banana", "Papaya", "Pineapple", "Watermelon",
    "Muskmelon", "Strawberry", "Blueberry", "Coffee", "Tea", "Rubber", "Coconut"
]

categories = ["Disease", "Crop Info", "Irrigation", "Soil", "Fertilizer", "Pests", "Weather", "Growth Stages", "General"]

severities = ["Low", "Medium", "High", "Critical"]

# Question templates by category
disease_templates = [
    ("Why are {crop} leaves yellow?", "{crop} leaves turning yellow can indicate nutrient deficiency, water stress, or fungal infection. Inspect for spots, wilting, or pest damage.", "Check soil moisture, inspect for disease symptoms, and consider foliar testing."),
    ("What causes brown spots on {crop} leaves?", "Brown spots on {crop} are commonly caused by fungal diseases like early blight or septoria leaf spot.", "Remove infected leaves, improve air circulation, and apply recommended fungicide if needed."),
    ("Why are {crop} leaves curling?", "Leaf curl in {crop} may be caused by aphids, viral infection, or herbicide drift.", "Inspect undersides for pests, check for virus symptoms, and avoid herbicide contamination."),
    ("What is causing {crop} to wilt?", "Wilting in {crop} can be due to water stress, fusarium wilt, or nematode damage.", "Check soil moisture, inspect roots for rot, and ensure proper drainage."),
    ("Why is {crop} showing white powder on leaves?", "White powdery growth on {crop} leaves is characteristic of powdery mildew.", "Improve air circulation, avoid overhead watering, and apply sulfur-based fungicide."),
    ("What causes black spots on {crop} fruit?", "Black spots on {crop} fruit are often caused by anthracnose or bacterial spot.", "Harvest ripe fruits promptly, remove infected ones, and apply copper-based bactericide."),
    ("Why are {crop} fruits dropping prematurely?", "Premature fruit drop in {crop} can result from water stress, nutrient deficiency, or pest attack.", "Maintain consistent irrigation, apply balanced fertilizer, and monitor for fruit borers."),
    ("What causes stem rot in {crop}?", "Stem rot in {crop} is typically caused by fungal pathogens in waterlogged conditions.", "Improve drainage, avoid overwatering, and remove affected plants."),
    ("Why are {crop} roots rotting?", "Root rot in {crop} is usually caused by pythium or phytophthora in poorly drained soil.", "Improve soil drainage, reduce irrigation frequency, and treat soil with fungicide if needed."),
    ("What causes mosaic patterns on {crop} leaves?", "Mosaic patterns indicate viral infection in {crop}, often transmitted by aphids or mechanical contact.", "Control aphid vectors, remove infected plants, and disinfect tools."),
    ("Why is {crop} growth stunted?", "Stunted growth in {crop} can result from poor soil fertility, root damage, or disease.", "Test soil nutrients, check for root problems, and ensure proper spacing."),
    ("What causes leaf spots in {crop}?", "Leaf spots in {crop} are mainly fungal or bacterial, favored by humid conditions.", "Apply preventive fungicide, ensure good air flow, and avoid wetting foliage."),
    ("Why are {crop} flowers falling off?", "Flower drop in {crop} may be due to water stress, temperature extremes, or poor pollination.", "Maintain consistent moisture, avoid temperature extremes, and encourage pollinators."),
    ("What causes rust on {crop}?", "Rust pustules on {crop} are caused by rust fungi, appearing as orange-brown spots.", "Apply rust-specific fungicide, remove infected leaves, and improve air circulation."),
    ("Why is {crop} turning purple?", "Purple discoloration in {crop} often indicates phosphorus deficiency or cold stress.", "Apply phosphorus-rich fertilizer, protect from cold, and check soil pH."),
]

crop_info_templates = [
    ("What is the best season to grow {crop}?", "{crop} is typically grown in {season} when temperatures range from {temp_range}.", "Prepare land during {prep_season}, use certified seeds, and follow recommended spacing."),
    ("How long does {crop} take to mature?", "{crop} takes approximately {days} days from sowing to harvest under normal conditions.", "Monitor crop development stages and harvest when mature for best yield."),
    ("What soil type is best for {crop}?", "{crop} grows best in {soil_type} with pH {ph_range}.", "Test soil before planting, amend with organic matter if needed, and ensure good drainage."),
    ("What is the ideal spacing for {crop}?", "Ideal spacing for {crop} is {spacing} to allow adequate light and air circulation.", "Follow recommended plant population for your variety and irrigation method."),
    ("How much water does {crop} need?", "{crop} requires approximately {water} mm of water per growing season.", "Irrigate based on soil moisture and crop stage; avoid waterlogging."),
    ("What are the common varieties of {crop}?", "Popular {crop} varieties include {varieties} suitable for different regions and seasons.", "Choose varieties based on local adaptability, disease resistance, and market demand."),
    ("How do I improve {crop} yield?", "To improve {crop} yield: use quality seeds, proper spacing, timely irrigation, balanced fertilization, and pest management.", "Adopt integrated crop management practices and monitor regularly."),
    ("What pests commonly attack {crop}?", "Common pests of {crop} include {pests}.", "Use IPM strategies: crop rotation, resistant varieties, biological control, and targeted pesticides."),
    ("What diseases affect {crop}?", "Major diseases of {crop} include {diseases}.", "Use disease-free seeds, practice crop rotation, and apply preventive fungicides."),
    ("Can {crop} be grown in pots?", "Yes, {crop} can be grown in pots with proper drainage, quality potting mix, and regular care.", "Use large containers, provide support, and maintain consistent watering."),
]

irrigation_templates = [
    ("How often should I irrigate {crop}?", "Irrigate {crop} every {interval} depending on soil type and weather conditions.", "Check soil moisture before irrigating; avoid frequent light irrigation."),
    ("What is the best irrigation method for {crop}?", "Drip irrigation is most efficient for {crop}, followed by sprinkler and furrow methods.", "Choose method based on water availability, soil type, and crop stage."),
    ("How much water does {crop} need per day?", "{crop} typically needs {water} mm per day during peak growth.", "Adjust based on evapotranspiration and rainfall."),
    ("When should I stop irrigating {crop} before harvest?", "Stop irrigating {crop} {days} days before harvest to improve quality and shelf life.", "Reduce irrigation gradually during ripening stage."),
    ("Is overwatering harmful for {crop}?", "Yes, overwatering can cause root rot and reduce yield in {crop}.", "Ensure proper drainage and water only when needed."),
    ("What are the signs of water stress in {crop}?", "Signs include wilting, leaf rolling, stunted growth, and premature flowering.", "Adjust irrigation schedule and monitor soil moisture regularly."),
]

soil_templates = [
    ("What is the best soil pH for {crop}?", "Optimal pH for {crop} is between {ph_min} and {ph_max}.", "Test soil regularly and adjust with lime or sulfur as needed."),
    ("How can I improve soil fertility for {crop}?", "Add organic manure, practice crop rotation, use cover crops, and apply balanced fertilizers.", "Conduct soil tests annually and follow nutrient management plans."),
    ("What soil type is suitable for {crop}?", "{crop} grows well in {soil_type}.", "Amend soil with compost if needed and ensure good drainage."),
    ("How do I test soil for {crop}?", "Collect soil samples from root zone, send to lab for analysis, and interpret results for fertilizer application.", "Test soil before planting and every 2-3 years thereafter."),
    ("Why is my soil hard for {crop}?", "Hard soil may be due to low organic matter, compaction, or high clay content.", "Add organic matter, avoid working wet soil, and practice deep tillage."),
]

fertilizer_templates = [
    ("What fertilizer is best for {crop}?", "Use balanced NPK fertilizer with higher {nutrient} during {stage} stage for {crop}.", "Apply based on soil test recommendations and crop requirements."),
    ("How much nitrogen does {crop} need?", "{crop} typically requires {n_kg} kg nitrogen per hectare.", "Split nitrogen application: basal + top-dressing for better efficiency."),
    ("When should I fertilize {crop}?", "Apply fertilizer at planting, then at {stage} stage for optimal uptake.", "Use split doses and place fertilizer near root zone."),
    ("What are the signs of nutrient deficiency in {crop}?", "Yellowing leaves indicate nitrogen deficiency, purple leaves suggest phosphorus deficiency, and brown edges suggest potassium deficiency.", "Apply deficient nutrients through soil or foliar application."),
    ("Can I use organic fertilizer for {crop}?", "Yes, compost, vermicompost, and well-rotted manure are excellent for {crop}.", "Apply organic manure 2-3 weeks before planting."),
]

pest_templates = [
    ("What are common pests of {crop}?", "Common pests include {pests}.", "Use IPM: monitor fields, use traps, introduce natural enemies, and apply pesticides only when needed."),
    ("How do I control aphids on {crop}?", "Use neem oil, insecticidal soap, or introduce ladybugs for biological control.", "Monitor early and treat before populations explode."),
    ("What is eating my {crop} leaves?", "Likely caterpillars, beetles, or locusts. Inspect at night for cutworms.", "Handpick if few, apply Bacillus thuringiensis, or use targeted insecticide."),
    ("How do I prevent whitefly on {crop}?", "Use yellow sticky traps, reflective mulches, and neem-based sprays.", "Control weeds and alternate hosts near the field."),
    ("What causes holes in {crop} leaves?", "Holes are typically caused by chewing insects like caterpillars or beetles.", "Identify the pest and apply appropriate control measures."),
]

weather_templates = [
    ("How does temperature affect {crop}?", "{crop} grows best between {min_temp}C and {max_temp}C.", "Use shade nets in summer and protect from frost in winter."),
    ("What is the effect of heavy rain on {crop}?", "Heavy rain can cause waterlogging, root rot, and fruit splitting in {crop}.", "Ensure drainage, stake plants, and harvest mature fruits before storms."),
    ("How does humidity affect {crop} disease?", "High humidity promotes fungal diseases in {crop}, especially leaf spots and blights.", "Improve spacing, prune for air flow, and apply preventive fungicides."),
    ("What to do during drought for {crop}?", "During drought, use mulching, drip irrigation, and drought-tolerant varieties for {crop}.", "Conserve water, reduce plant population if needed, and monitor for stress."),
    ("How does frost affect {crop}?", "Frost damages {crop} tissues, causing cell rupture and tissue death.", "Use frost cloth, heaters, or anti-transpirants; irrigate before expected frost."),
]

growth_stage_templates = [
    ("What is the germination period for {crop}?", "{crop} germinates in {days} days under optimal conditions.", "Ensure proper seed depth, moisture, and temperature for uniform germination."),
    ("When should I transplant {crop}?", "Transplant {crop} when seedlings have {leaves} true leaves and are {height} cm tall.", "Harden seedlings before transplanting and irrigate immediately after."),
    ("What is the flowering stage of {crop}?", "{crop} flowers at {days} days after planting, lasting for {duration} days.", "Avoid water stress during flowering and ensure pollination."),
    ("When is {crop} ready to harvest?", "{crop} is ready to harvest at {days} days, when fruits/grains reach maturity.", "Harvest at proper maturity for best quality and storage."),
    ("What care does {crop} need during fruiting?", "During fruiting, {crop} needs consistent moisture and additional potassium.", "Support heavy fruits, monitor for pests, and avoid water stress."),
]

general_templates = [
    ("How do I start organic farming with {crop}?", "Start with organic seeds, compost, neem-based pesticides, and crop rotation for {crop}.", "Get organic certification, maintain soil health, and keep records."),
    ("What is crop rotation for {crop}?", "Rotate {crop} with non-host crops like legumes or cereals to break pest cycles.", "Plan a 3-4 year rotation cycle for sustainable production."),
    ("How do I control weeds in {crop}?", "Use mulching, manual weeding, or herbicides as last resort for {crop}.", "Weed early and regularly; use pre-emergence herbicides if needed."),
    ("What are the benefits of mulching for {crop}?", "Mulching conserves moisture, suppresses weeds, and moderates soil temperature for {crop}.", "Use straw, plastic mulch, or living mulches depending on crop."),
    ("How do I store {crop} after harvest?", "Dry {crop} to safe moisture levels, store in clean containers, and protect from pests.", "Use hermetic storage bags or silos for long-term storage."),
]

# Season/temp/soil etc helpers
seasons = ["Kharif", "Rabi", "Summer", "Zaid"]
temp_ranges = ["25-35C", "20-30C", "15-25C", "30-40C"]
soil_types = ["loamy soil", "sandy loam", "clay loam", "alluvial soil", "black soil", "red soil"]
ph_ranges = ["6.0-7.0", "6.5-7.5", "5.5-6.5", "7.0-8.0"]
spacings = ["30x30 cm", "45x45 cm", "60x60 cm", "15x15 cm", "90x90 cm"]
water_reqs = ["500-800", "800-1200", "300-500", "1000-1500"]
days_map = {
    "Pigeon Pea": 150, "Rice": 120, "Wheat": 120, "Maize": 100, "Cotton": 180,
    "Tomato": 90, "Potato": 90, "Onion": 120, "Soybean": 100, "Groundnut": 120,
    "Chickpea": 100, "Sugarcane": 360, "Millet": 90, "Sorghum": 100, "Banana": 300,
    "Mango": 150, "Grapes": 150, "Chilli": 120, "Brinjal": 120, "Okra": 60,
    "Cabbage": 90, "Cauliflower": 90, "Cucumber": 60, "Peas": 80, "Apple": 150, "Papaya": 300, "Pineapple": 400, "Watermelon": 90, "Muskmelon": 90, "Strawberry": 120, "Blueberry": 200, "Coffee": 400, "Tea": 300, "Rubber": 400, "Coconut": 300
}
varieties_map = {
    "Pigeon Pea": "ICPL 88039, BDN 2, GT 100", "Rice": "IR 64, MTU 1010, CR Dhan 201", "Wheat": "HD 2967, PBW 343, DBW 17", "Maize": "Hybrid 9022, PAC 740, Vivek 9", "Cotton": "Bunny Bt, RCH 2 Bt, Suraj Bt", "Tomato": "Arka Rakshak, Pusa Rohini, Hybrid 1", "Potato": "Kufri Jyoti, Kufri Chipsona, Atlantic", "Onion": "AgriFound Dark Red, N 53, Arka Niketan", "Soybean": "JS 335, MAUS 2, JS 95-60", "Groundnut": "TG 37A, JL 24, GG 20", "Chickpea": "JG 11, Vijay, ICCV 10", "Sugarcane": "Co 86032, Co 0238, Co 06030", "Millet": "HMR 1, RHRBI 60, ICTP 8203", "Sorghum": "CSH 23, SPV 462, M 35-1", "Banana": "Grand Naine, Robusta, Dwarf Cavendish", "Mango": "Alphonso, Kesar, Dashehari", "Grapes": "Thompson Seedless, Sonaka, Sharad Seedless", "Chilli": "Jwala, Kashi Anmol, Hybrid 1", "Brinjal": "Pusa Purple, Arka Shirish, Hybrid 1", "Okra": "Arka Anamika, Pusa Sawani, Hybrid 1", "Cabbage": "Golden Acre, Pride of India, Hybrid 1", "Cauliflower": "Snowball 16, Pusa Snowball, Hybrid 1", "Cucumber": "Pusa Uday, Hybrid 1, Poinsett 76", "Peas": "Arkel, Pusa Pragati, Hybrid 1", "Apple": "Red Delicious, Gala, Fuji", "Papaya": "Pusa Delicious, Red Lady, Taiwan", "Pineapple": "Queen, Kew, MD-2", "Watermelon": "Sugar Baby, Arka Manik, Hybrid 1", "Muskmelon": "Arka Rajhans, Durgapura Madhu, Hybrid 1", "Strawberry": "Sweet Charlie, Festival, Camarosa", "Blueberry": "Duke, Bluecrop, Legacy", "Coffee": "Arabica, Robusta, Liberica", "Tea": "Assamica, Sinensis, Camellia", "Rubber": "RRIM 600, RRII 105, PB 260", "Coconut": "West Coast Tall, East Coast Tall, Chowghat Orange Dwarf"
}
pests_map = {
    "Pigeon Pea": "pod borer, pod fly, aphids", "Rice": "brown planthopper, stem borer, leaf folder", "Wheat": "aphids, termites, armyworm", "Maize": "fall armyworm, stem borer, shoot fly", "Cotton": "bollworm, whitefly, jassid", "Tomato": "fruit borer, whitefly, leaf miner", "Potato": "tuber moth, aphids, Colorado beetle", "Onion": "thrips, maggot, onion fly", "Soybean": "girdle beetle, aphids, whitefly", "Groundnut": "tikka leaf spot, aphids, jassid", "Chickpea": "pod borer, aphids, cutworm", "Sugarcane": "borer, scale insect, pyrilla", "Millet": "shoot fly, stem borer, aphids", "Sorghum": "shoot fly, stem borer, aphids", "Banana": "weevil borer, aphids, nematode", "Mango": "mango hopper, fruit fly, stone weevil", "Grapes": "thrips, mealybug, flea beetle", "Chilli": "mites, aphids, thrips", "Brinjal": "fruit borer, aphids, whitefly", "Okra": "fruit borer, aphids, jassid", "Cabbage": "diamondback moth, aphids, cutworm", "Cauliflower": "diamondback moth, aphids, leaf webber", "Cucumber": "fruit fly, aphids, red pumpkin beetle", "Peas": "pea weevil, aphids, pod borer", "Apple": "codling moth, aphids,San Jose scale", "Papaya": "fruit fly, aphids, leaf curl", "Pineapple": "mealybug, scale insect, pineapple wilt", "Watermelon": "fruit fly, aphids, thrips", "Muskmelon": "fruit fly, aphids, thrips", "Strawberry": "aphids, mites, crown rot", "Blueberry": "mites, fruit worms, birds", "Coffee": "berry borer, white stem borer, leaf rust", "Tea": "mites, thrips, tea mosquito bug", "Rubber": "cockchafer, termites, leaf fall disease", "Coconut": "rhinoceros beetle, red palm weevil, mite"
}
diseases_map = {
    "Pigeon Pea": "wilt, sterility mosaic, yellow mosaic", "Rice": "blast, brown spot, bacterial leaf blight", "Wheat": "rust, powdery mildew, loose smut", "Maize": "leaf blight, downy mildew, stalk rot", "Cotton": "bacterial blight, wilt, leaf spot", "Tomato": "early blight, late blight, spotted wilt", "Potato": "early blight, late blight, blackleg", "Onion": "purple blotch, downy mildew, basal rot", "Soybean": "rust, bacterial blight, mosaic", "Groundnut": "tikka leaf spot, rust, wilt", "Chickpea": "wilt, blight, grey mold", "Sugarcane": "smut, red rot, wilt", "Millet": "downy mildew, smut, blast", "Sorghum": "smut, downy mildew, anthracnose", "Banana": "Panama wilt, bunchy top, Sigatoka", "Mango": "powdery mildew, anthracnose, sooty mold", "Grapes": "downy mildew, powdery mildew, anthracnose", "Chilli": "leaf curl, powdery mildew, damping off", "Brinjal": "phomopsis blight, bacterial wilt, little leaf", "Okra": "yellow vein mosaic, powdery mildew, wilt", "Cabbage": "black rot, clubroot, downy mildew", "Cauliflower": "black rot, clubroot, downy mildew", "Cucumber": "powdery mildew, downy mildew, wilt", "Peas": "powdery mildew, wilt, rust", "Apple": "scab, fire blight, cedar apple rust", "Papaya": "mosaic virus, powdery mildew, anthracnose", "Pineapple": "heart rot, root rot, pineapple wilt", "Watermelon": "powdery mildew, downy mildew, anthracnose", "Muskmelon": "powdery mildew, downy mildew, fusarium wilt", "Strawberry": "leaf spot, powdery mildew, gray mold", "Blueberry": "mummy berry, anthracnose, botrytis", "Coffee": "leaf rust, berry disease, bacterial blight", "Tea": "blight, root rot, anthracnose", "Rubber": "powdery mildew, Colletotrichum, Phytophthora", "Coconut": "lethal yellowing, bud rot, stem bleeding"
}

def get_season(crop):
    mapping = {
        "Pigeon Pea": "Kharif", "Rice": "Kharif/Rabi", "Wheat": "Rabi", "Maize": "Kharif/Rabi",
        "Cotton": "Kharif", "Tomato": "Rabi/Summer", "Potato": "Rabi", "Onion": "Rabi",
        "Soybean": "Kharif", "Groundnut": "Kharif", "Chickpea": "Rabi", "Sugarcane": "Annual",
        "Millet": "Kharif", "Sorghum": "Kharif", "Banana": "Perennial", "Mango": "Perennial",
        "Grapes": "Perennial", "Chilli": "Kharif/Rabi", "Brinjal": "Rabi/Summer", "Okra": "Kharif/Summer",
        "Cabbage": "Rabi", "Cauliflower": "Rabi", "Cucumber": "Summer", "Peas": "Rabi",
        "Apple": "Temperate", "Papaya": "Perennial", "Pineapple": "Perennial", "Watermelon": "Summer",
        "Muskmelon": "Summer", "Strawberry": "Temperate", "Blueberry": "Temperate", "Coffee": "Tropical",
        "Tea": "Tropical", "Rubber": "Tropical", "Coconut": "Tropical"
    }
    return mapping.get(crop, "Kharif")

entry_id = 1
entries = []

# Generate disease entries
for crop in crops[:20]:
    for q, ans, rec in disease_templates:
        topic = q.replace("Why are", "").replace("What causes", "").replace("{crop}", "").strip().title()
        question = q.format(crop=crop)
        keywords = f"{crop}, {topic.replace(' ', ', ')}, disease, problem"
        answer = ans.format(crop=crop)
        recommendation = rec
        severity = random.choice(["Low", "Medium", "High", "Critical"])
        entries.append([entry_id, "Disease", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate crop info entries
for crop in crops[:25]:
    for q, ans, rec in crop_info_templates:
        topic = "Crop Information"
        question = q.format(crop=crop, season=get_season(crop), temp_range=random.choice(temp_ranges),
                           prep_season=random.choice(seasons), days=days_map.get(crop, 120),
                           soil_type=random.choice(soil_types), ph_range=random.choice(ph_ranges),
                           spacing=random.choice(spacings), water=random.choice(water_reqs),
                           varieties=varieties_map.get(crop, "local varieties"),
                           pests=pests_map.get(crop, "various pests"),
                           diseases=diseases_map.get(crop, "various diseases"))
        keywords = f"{crop}, cultivation, growing, farming, {topic.lower()}"
        answer = ans.format(crop=crop, season=get_season(crop), temp_range=random.choice(temp_ranges),
                           days=days_map.get(crop, 120), soil_type=random.choice(soil_types),
                           ph_range=random.choice(ph_ranges), spacing=random.choice(spacings),
                           water=random.choice(water_reqs), varieties=varieties_map.get(crop, "local varieties"),
                           pests=pests_map.get(crop, "various pests"),
                           diseases=diseases_map.get(crop, "various diseases"))
        recommendation = rec.format(crop=crop, prep_season=random.choice(seasons))
        severity = "Low"
        entries.append([entry_id, "Crop Info", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate irrigation entries
for crop in crops[:25]:
    for q, ans, rec in irrigation_templates:
        topic = "Irrigation"
        question = q.format(crop=crop, interval=random.choice(["7-10 days", "10-15 days", "5-7 days"]))
        keywords = f"{crop}, irrigation, water, watering, moisture"
        answer = ans.format(crop=crop, interval=random.choice(["7-10 days", "10-15 days", "5-7 days"]),
                           water=random.choice(["4-6", "6-8", "3-5"]), days=random.randint(3,10))
        recommendation = rec.format(crop=crop)
        severity = "Low"
        entries.append([entry_id, "Irrigation", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate soil entries
for crop in crops[:25]:
    for q, ans, rec in soil_templates:
        topic = "Soil"
        question = q.format(crop=crop, ph_min=random.choice(["5.5", "6.0", "6.5"]), ph_max=random.choice(["7.0", "7.5", "8.0"]))
        keywords = f"{crop}, soil, fertility, pH, nutrients"
        answer = ans.format(crop=crop, ph_min=random.choice(["5.5", "6.0", "6.5"]), ph_max=random.choice(["7.0", "7.5", "8.0"]), soil_type=random.choice(soil_types))
        recommendation = rec.format(crop=crop)
        severity = "Low"
        entries.append([entry_id, "Soil", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate fertilizer entries
for crop in crops[:25]:
    for q, ans, rec in fertilizer_templates:
        topic = "Fertilizer"
        question = q.format(crop=crop, nutrient=random.choice(["potassium", "phosphorus", "nitrogen"]), stage=random.choice(["vegetative", "flowering", "fruiting"]), n_kg=random.randint(50, 150))
        keywords = f"{crop}, fertilizer, nutrients, NPK, manure"
        answer = ans.format(crop=crop, nutrient=random.choice(["potassium", "phosphorus", "nitrogen"]), stage=random.choice(["vegetative", "flowering", "fruiting"]), n_kg=random.randint(50, 150))
        recommendation = rec.format(crop=crop)
        severity = "Low"
        entries.append([entry_id, "Fertilizer", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate pest entries
for crop in crops[:25]:
    for q, ans, rec in pest_templates:
        topic = "Pests"
        question = q.format(crop=crop, pests=pests_map.get(crop, "common pests"))
        keywords = f"{crop}, pests, insects, control, IPM, {pests_map.get(crop, 'pests')}"
        answer = ans.format(crop=crop, pests=pests_map.get(crop, "common pests"))
        recommendation = rec
        severity = random.choice(["Low", "Medium", "High"])
        entries.append([entry_id, "Pests", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate weather entries
for crop in crops[:25]:
    for q, ans, rec in weather_templates:
        topic = "Weather"
        question = q.format(crop=crop, min_temp=random.randint(15,25), max_temp=random.randint(30,40), days=random.randint(5,15))
        keywords = f"{crop}, weather, temperature, rain, humidity, stress"
        answer = ans.format(crop=crop, min_temp=random.randint(15,25), max_temp=random.randint(30,40), days=random.randint(5,15))
        recommendation = rec.format(crop=crop)
        severity = random.choice(["Low", "Medium", "High"])
        entries.append([entry_id, "Weather", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate growth stage entries
for crop in crops[:25]:
    for q, ans, rec in growth_stage_templates:
        topic = "Growth Stages"
        question = q.format(crop=crop, days=days_map.get(crop, 120), leaves=random.randint(3,6), height=random.randint(10,20), duration=random.randint(20,60))
        keywords = f"{crop}, growth, stages, germination, flowering, harvest"
        answer = ans.format(crop=crop, days=days_map.get(crop, 120), leaves=random.randint(3,6), height=random.randint(10,20), duration=random.randint(20,60))
        recommendation = rec.format(crop=crop)
        severity = "Low"
        entries.append([entry_id, "Growth Stages", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

# Generate general entries
for crop in crops[:25]:
    for q, ans, rec in general_templates:
        topic = "General"
        question = q.format(crop=crop)
        keywords = f"{crop}, general, farming, organic, management"
        answer = ans.format(crop=crop)
        recommendation = rec.format(crop=crop)
        severity = "Low"
        entries.append([entry_id, "General", crop, topic, question, keywords, answer, recommendation, severity, "English"])
        entry_id += 1

print(f"Total entries generated: {len(entries)}")
for row in entries[:5]:
    ws.append(row)

for row in entries[5:]:
    ws.append(row)

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUTPUT)
print(f"Saved to {OUTPUT}")
print(f"Total rows including header: {ws.max_row}")
