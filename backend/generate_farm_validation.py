import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "farm_validation.xlsx")

HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0'),
)


def style_header(ws, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def add_sheet_data(ws, headers, rows):
    ws.append(headers)
    style_header(ws, len(headers))
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = ALT_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = Alignment(vertical='top', wrap_text=True)
    auto_width(ws)


wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: Crop-Soil Compatibility Rules
ws_rules = wb.create_sheet("Crop Soil Rules")
rules_headers = [
    "Crop", "Soil Type", "Suitability", "Explanation",
    "Recommended Action", "Amendments Required", "Irrigation Adjustment"
]
rules_data = [
    ["Rice", "Black soil", "LOW",
     "Black soil has high water retention but poor drainage. Rice requires standing water but black soil can become waterlogged, causing root rot and nutrient deficiency.",
     "Avoid direct rice cultivation in pure black soil. Consider raised beds or alternate crops like cotton or soybean. If rice is necessary, ensure proper drainage channels.",
     "Add sand to improve drainage. Apply gypsum to reduce alkalinity. Incorporate organic compost.",
     "Use alternate wetting and drying instead of continuous flooding. Ensure field drains every 7-10 days."],
    ["Rice", "Red soil", "LOW",
     "Red soil drains quickly and lacks water retention needed for rice. Low fertility and acidity make it unsuitable for rice without significant amendments.",
     "Use drought-tolerant rice varieties or consider alternative crops like millets or pulses. If rice is necessary, use System of Rice Intensification (SRI).",
     "Add organic matter and clay to improve water retention. Apply lime to correct acidity. Use balanced NPK fertilizers.",
     "Use frequent light irrigation. Consider drip irrigation with water harvesting."],
    ["Rice", "Alluvial", "HIGH",
     "Alluvial soil is ideal for rice with excellent water retention, fertility, and drainage. Rich in organic matter and nutrients.",
     "No major amendments needed. Maintain proper water levels. Apply nitrogen in split doses for optimal yield.",
     "Apply zinc if deficient. Maintain organic matter through crop rotation.",
     "Maintain 5cm standing water. Use alternate wetting and drying for water conservation."],
    ["Rice", "Sandy", "LOW",
     "Sandy soil drains too quickly and cannot hold standing water needed for rice. Very low nutrient retention.",
     "Consider System of Rice Intensification (SRI) with minimal water. Alternatively, grow drought-tolerant crops like millets.",
     "Add heavy clay or organic compost. Use mulching to reduce evaporation.",
     "Use very frequent light irrigation. Consider laser land leveling for uniform water distribution."],
    ["Rice", "Clay", "MODERATE",
     "Clay soil holds water well but can become waterlogged. Needs proper management to prevent anaerobic conditions.",
     "Ensure proper puddling before transplanting. Use drainage channels. Monitor for soil salinity.",
     "Apply organic matter to improve structure. Use gypsum if saline.",
     "Maintain shallow water levels. Use periodic drainage to aerate soil."],
    ["Rice", "Loamy", "HIGH",
     "Loamy soil provides ideal conditions for rice with good drainage, water retention, and fertility.",
     "No major amendments needed. Apply nitrogen in split doses. Ensure proper seed bed preparation.",
     "Maintain organic matter through green manuring.",
     "Maintain 5cm standing water. Use SRI methods for water efficiency."],

    ["Wheat", "Black soil", "MODERATE",
     "Black soil is good for wheat but can become hard and crack during dry periods. Retains moisture well but needs careful water management.",
     "Use drought-tolerant wheat varieties. Apply organic matter to improve soil structure. Monitor soil moisture regularly.",
     "Add gypsum to improve soil structure. Apply balanced NPK fertilizers.",
     "Use light but frequent irrigation. Avoid overwatering which causes waterlogging."],
    ["Wheat", "Red soil", "MODERATE",
     "Red soil drains well but has low fertility and water retention. Suitable for wheat with proper fertilization and irrigation.",
     "Apply phosphorus and potassium fertilizers. Use organic compost. Consider liming if acidic.",
     "Add organic matter. Apply zinc sulfate if deficient.",
     "Use frequent light irrigation. Apply irrigation at crown root initiation and milking stages."],
    ["Wheat", "Alluvial", "HIGH",
     "Alluvial soil is highly suitable for wheat with good fertility, drainage, and water retention. Ideal for Rabi season.",
     "No major amendments needed. Apply balanced NPK fertilizers. Ensure timely sowing in October-November.",
     "Apply zinc if deficient. Maintain organic matter through crop rotation.",
     "Use 4-5 irrigations at critical stages: crown root, tillering, milking, dough, and late grain filling."],
    ["Wheat", "Sandy", "LOW",
     "Sandy soil drains too quickly and has low nutrient retention. Wheat can grow but yields will be low without amendments.",
     "Use drought-tolerant varieties. Apply heavy organic matter. Consider mixed cropping with legumes.",
     "Add clay or silt. Apply frequent small doses of fertilizers. Use mulching.",
     "Use very frequent light irrigation. Consider drip irrigation."],
    ["Wheat", "Clay", "MODERATE",
     "Clay soil retains water well but can become waterlogged in heavy rain. Needs good drainage for wheat.",
     "Ensure proper drainage. Use raised beds in waterlogged areas. Apply organic matter to improve structure.",
     "Add gypsum to improve structure. Apply organic compost.",
     "Avoid overwatering. Use drainage channels. Irrigate at critical stages only."],
    ["Wheat", "Loamy", "HIGH",
     "Loamy soil is ideal for wheat with excellent drainage, fertility, and water holding capacity.",
     "No major amendments needed. Apply nitrogen in split doses. Ensure proper seed bed preparation.",
     "Maintain organic matter through crop rotation with legumes.",
     "Use 4-5 irrigations at critical stages. Avoid waterlogging."],

    ["Cotton", "Black soil", "HIGH",
     "Black soil is the best for cotton with excellent water retention and fertility. Deep cracks allow aeration during dry periods.",
     "No major amendments needed. Ensure proper drainage during monsoon. Apply balanced fertilizers.",
     "Apply gypsum if saline. Maintain organic matter.",
     "Use moderate irrigation. Avoid waterlogging. Ensure drainage during heavy rains."],
    ["Cotton", "Red soil", "MODERATE",
     "Red soil can support cotton with proper amendments. Has good drainage but low fertility and water retention.",
     "Add organic matter and phosphorus. Improve moisture retention through mulching.",
     "Apply organic compost. Use potassium-rich fertilizers.",
     "Use frequent light irrigation. Apply irrigation at square, boll, and boll maturation stages."],
    ["Cotton", "Alluvial", "MODERATE",
     "Alluvial soil is good for cotton but may need more frequent irrigation due to good drainage.",
     "Improve drainage. Apply potassium-rich fertilizer. Use organic matter.",
     "Add organic compost. Apply balanced NPK with extra potassium.",
     "Increase irrigation frequency. Monitor for boll rot during heavy rains."],
    ["Cotton", "Sandy", "LOW",
     "Sandy soil drains too quickly and lacks nutrients for cotton. Water stress will reduce yield significantly.",
     "Add heavy organic matter. Use drought-tolerant varieties. Consider alternative crops.",
     "Add clay or silt. Apply frequent small fertilizer doses.",
     "Use very frequent irrigation. Consider drip irrigation system."],
    ["Cotton", "Clay", "LOW",
     "Clay soil retains too much water and can become waterlogged, causing root rot and boll rot in cotton.",
     "Improve drainage significantly. Use raised beds. Consider alternative crops like rice.",
     "Add sand and organic matter. Apply gypsum.",
     "Ensure excellent drainage. Avoid irrigation during heavy rain periods."],
    ["Cotton", "Loamy", "HIGH",
     "Loamy soil is ideal for cotton with good drainage, fertility, and water holding capacity.",
     "No major amendments needed. Apply balanced NPK fertilizers. Ensure proper spacing.",
     "Maintain organic matter. Apply potassium for boll development.",
     "Use moderate irrigation. Apply at square, boll, and boll maturation stages."],

    ["Maize", "Black soil", "MODERATE",
     "Black soil retains moisture well but can become waterlogged. Good for maize with proper drainage.",
     "Ensure proper drainage. Use organic matter. Apply balanced fertilizers.",
     "Add gypsum if needed. Apply organic compost.",
     "Use moderate irrigation. Avoid waterlogging during early growth."],
    ["Maize", "Red soil", "MODERATE",
     "Red soil drains well but has low fertility. Suitable for maize with proper fertilization.",
     "Apply nitrogen-rich fertilizers. Use organic compost. Consider liming.",
     "Add organic matter. Apply zinc sulfate.",
     "Use frequent light irrigation. Apply at knee-high and tasseling stages."],
    ["Maize", "Alluvial", "HIGH",
     "Alluvial soil is excellent for maize with good fertility and drainage. Warm soil promotes rapid growth.",
     "No major amendments needed. Apply nitrogen at planting and side-dress at knee-high.",
     "Maintain organic matter through crop rotation.",
     "Use moderate irrigation. Apply at planting, knee-high, and tasseling stages."],
    ["Maize", "Sandy", "MODERATE",
     "Sandy soil warms quickly in spring, good for early maize. But drains quickly and needs frequent irrigation.",
     "Apply organic matter to improve water retention. Mulch to conserve moisture.",
     "Add clay or organic compost. Apply frequent small fertilizer doses.",
     "Use very frequent light irrigation. Consider drip irrigation."],
    ["Maize", "Clay", "LOW",
     "Clay soil can become waterlogged and compacted, inhibiting root growth in maize. Cold soil delays germination.",
     "Improve drainage. Use organic matter. Consider raised beds.",
     "Add sand and organic matter. Apply gypsum.",
     "Avoid overwatering. Use drainage channels. Plant on ridges."],
    ["Maize", "Loamy", "HIGH",
     "Loamy soil is ideal for maize with good drainage, fertility, and warm soil conditions.",
     "No major amendments needed. Apply nitrogen at planting and side-dress at knee-high.",
     "Maintain organic matter. Apply phosphorus for root development.",
     "Use moderate irrigation. Apply at planting, knee-high, and tasseling stages."],

    ["Soybean", "Black soil", "HIGH",
     "Black soil is excellent for soybean with good water retention and fertility. Rich in nutrients.",
     "No major amendments needed. Inoculate seeds with rhizobium. Apply phosphorus.",
     "Maintain organic matter. Apply potassium if deficient.",
     "Use moderate irrigation. Avoid waterlogging during germination."],
    ["Soybean", "Red soil", "MODERATE",
     "Red soil can support soybean with proper amendments. Has good drainage but low fertility.",
     "Apply phosphorus and potassium. Use organic compost. Inoculate seeds.",
     "Add organic matter. Apply rhizobium inoculation.",
     "Use moderate irrigation. Apply at flowering and pod filling stages."],
    ["Soybean", "Alluvial", "HIGH",
     "Alluvial soil is highly suitable for soybean with good fertility and drainage.",
     "No major amendments needed. Inoculate seeds. Apply phosphorus and potassium.",
     "Maintain organic matter through crop rotation.",
     "Use moderate irrigation. Apply at flowering and pod filling stages."],
    ["Soybean", "Sandy", "LOW",
     "Sandy soil drains too quickly and has low nutrient retention. Poor for soybean.",
     "Add heavy organic matter. Use drought-tolerant varieties. Consider alternative crops.",
     "Add clay or silt. Apply frequent small fertilizer doses.",
     "Use very frequent irrigation. Consider drip irrigation."],
    ["Soybean", "Clay", "MODERATE",
     "Clay soil retains moisture well but can become waterlogged. Good for soybean with proper drainage.",
     "Ensure good drainage. Avoid waterlogging during germination. Use organic matter.",
     "Add organic compost. Apply gypsum if saline.",
     "Avoid overwatering. Use drainage channels."],
    ["Soybean", "Loamy", "HIGH",
     "Loamy soil is ideal for soybean with good drainage, fertility, and moisture retention.",
     "No major amendments needed. Inoculate seeds with rhizobium. Apply phosphorus.",
     "Maintain organic matter. Apply potassium for pod filling.",
     "Use moderate irrigation. Apply at flowering and pod filling stages."],

    ["Sugarcane", "Black soil", "HIGH",
     "Black soil is excellent for sugarcane with high water retention and fertility. Deep soil allows extensive root growth.",
     "No major amendments needed. Apply high potassium. Ensure consistent irrigation.",
     "Maintain organic matter. Apply zinc if deficient.",
     "Use frequent irrigation. Maintain consistent soil moisture. Avoid waterlogging."],
    ["Sugarcane", "Red soil", "MODERATE",
     "Red soil can support sugarcane with amendments. Has good drainage but low fertility and water retention.",
     "Apply high potassium and nitrogen. Use organic compost. Improve moisture retention.",
     "Add organic matter. Apply heavy NPK fertilizers.",
     "Use very frequent irrigation. Consider drip irrigation system."],
    ["Sugarcane", "Alluvial", "HIGH",
     "Alluvial soil is highly suitable for sugarcane with high fertility and water holding capacity.",
     "No major amendments needed. Apply nitrogen in split doses. Ensure proper drainage.",
     "Maintain organic matter. Apply potassium for sugar accumulation.",
     "Use frequent irrigation. Maintain consistent soil moisture."],
    ["Sugarcane", "Sandy", "LOW",
     "Sandy soil drains too quickly and has low nutrient retention. Very challenging for sugarcane.",
     "Add heavy organic matter. Use drip irrigation. Consider alternative crops.",
     "Add clay or silt. Apply frequent small fertilizer doses.",
     "Use very frequent irrigation. Drip irrigation essential."],
    ["Sugarcane", "Clay", "HIGH",
     "Clay soil retains water well and is fertile. Good for sugarcane but needs drainage management.",
     "Ensure proper drainage. Apply organic matter. Use balanced fertilizers.",
     "Add organic compost. Apply gypsum if saline.",
     "Use frequent irrigation. Avoid waterlogging. Ensure drainage channels."],
    ["Sugarcane", "Loamy", "HIGH",
     "Loamy soil is ideal for sugarcane with good drainage, fertility, and moisture retention.",
     "No major amendments needed. Apply nitrogen in split doses. Ensure consistent irrigation.",
     "Maintain organic matter. Apply potassium for sugar content.",
     "Use frequent irrigation. Maintain consistent soil moisture."],

    ["Groundnut", "Black soil", "MODERATE",
     "Black soil retains moisture well but can become waterlogged. Groundnut needs well-drained soil for pod development.",
     "Ensure proper drainage. Use raised beds. Apply gypsum at pegging stage.",
     "Add sand to improve drainage. Apply gypsum.",
     "Avoid overwatering. Use light irrigation. Ensure good drainage."],
    ["Groundnut", "Red soil", "HIGH",
     "Red soil with good drainage is ideal for groundnut. Sandy loam texture allows pod development.",
     "No major amendments needed. Apply gypsum at pegging stage. Use balanced fertilizers.",
     "Apply gypsum at pegging. Maintain organic matter.",
     "Use light irrigation. Avoid waterlogging. Apply at pegging and pod filling stages."],
    ["Groundnut", "Alluvial", "MODERATE",
     "Alluvial soil is good for groundnut but may need sand addition for better drainage.",
     "Add sand if soil is heavy. Apply gypsum. Use balanced fertilizers.",
     "Add sand if needed. Apply gypsum at pegging.",
     "Use light irrigation. Avoid waterlogging during pod development."],
    ["Groundnut", "Sandy", "HIGH",
     "Sandy loam soil is ideal for groundnut pod development. Loose texture allows pegging and harvesting.",
     "No major amendments needed. Apply gypsum at pegging. Use balanced fertilizers.",
     "Add organic matter. Apply gypsum at pegging stage.",
     "Use light but frequent irrigation. Avoid waterlogging."],
    ["Groundnut", "Clay", "LOW",
     "Clay soil is too heavy for groundnut. Pods cannot develop in compacted soil. High water retention causes pod rot.",
     "Add sand to improve texture. Use raised beds. Consider alternative crops.",
     "Add sand and organic matter. Apply gypsum.",
     "Avoid overwatering. Ensure excellent drainage."],
    ["Groundnut", "Loamy", "HIGH",
     "Loamy soil is ideal for groundnut with good drainage and loose texture for pod development.",
     "No major amendments needed. Apply gypsum at pegging. Use balanced fertilizers.",
     "Add organic matter. Apply gypsum at pegging.",
     "Use light irrigation. Avoid waterlogging during pod development."],

    ["Chickpea", "Black soil", "MODERATE",
     "Black soil retains moisture well but can become waterlogged. Chickpea needs well-drained soil.",
     "Ensure proper drainage. Use raised beds. Treat seeds with fungicide.",
     "Add sand if heavy. Apply organic matter.",
     "Avoid overwatering. Use light irrigation. Ensure good drainage."],
    ["Chickpea", "Red soil", "HIGH",
     "Red soil with good drainage is ideal for chickpea. Well-drained sandy loam is best.",
     "No major amendments needed. Treat seeds with fungicide. Apply rhizobium inoculation.",
     "Add organic matter. Apply phosphorus.",
     "Use very light irrigation. Chickpea is drought-tolerant. Avoid waterlogging."],
    ["Chickpea", "Alluvial", "MODERATE",
     "Alluvial soil is good for chickpea but may need sand addition for better drainage.",
     "Add sand if soil is heavy. Treat seeds with fungicide. Apply balanced fertilizers.",
     "Add sand if needed. Apply rhizobium inoculation.",
     "Use light irrigation. Avoid waterlogging."],
    ["Chickpea", "Sandy", "HIGH",
     "Sandy loam soil is ideal for chickpea with excellent drainage. Drought-tolerant crop.",
     "No major amendments needed. Treat seeds with fungicide. Apply balanced fertilizers.",
     "Add organic matter. Apply phosphorus.",
     "Use minimal irrigation. Rainfed cultivation suitable. Avoid overwatering."],
    ["Chickpea", "Clay", "LOW",
     "Clay soil is too heavy and retains too much water. Causes root rot and wilt in chickpea.",
     "Add sand to improve drainage. Use raised beds. Consider alternative crops.",
     "Add sand and organic matter. Apply gypsum.",
     "Avoid waterlogging. Ensure excellent drainage."],
    ["Chickpea", "Loamy", "HIGH",
     "Loamy soil is ideal for chickpea with good drainage and fertility.",
     "No major amendments needed. Treat seeds with fungicide. Apply rhizobium inoculation.",
     "Add organic matter. Apply phosphorus.",
     "Use light irrigation. Rainfed cultivation suitable."],

    ["Pigeon Pea", "Black soil", "HIGH",
     "Black soil is ideal for pigeon pea with excellent water retention and fertility. Drought-resistant crop.",
     "No major amendments needed. Apply phosphorus. Use balanced fertilizers.",
     "Maintain organic matter. Apply zinc if deficient.",
     "Use rainfed or light irrigation. Drought-tolerant crop. Avoid waterlogging."],
    ["Pigeon Pea", "Red soil", "HIGH",
     "Red soil with good drainage is suitable for pigeon pea. Well-drained soil preferred.",
     "No major amendments needed. Apply phosphorus. Use balanced fertilizers.",
     "Add organic matter. Apply phosphorus.",
     "Use rainfed or light irrigation. Drought-tolerant crop."],
    ["Pigeon Pea", "Alluvial", "MODERATE",
     "Alluvial soil is good for pigeon pea but may need sand addition for better drainage.",
     "Add sand if soil is heavy. Apply phosphorus. Use balanced fertilizers.",
     "Add organic matter. Apply phosphorus.",
     "Use light irrigation. Drought-tolerant crop."],
    ["Pigeon Pea", "Sandy", "MODERATE",
     "Sandy soil drains well but has low nutrient retention. Suitable with organic amendments.",
     "Add heavy organic matter. Apply phosphorus. Use balanced fertilizers.",
     "Add organic compost. Apply phosphorus.",
     "Use light irrigation. Drought-tolerant crop."],
    ["Pigeon Pea", "Clay", "LOW",
     "Clay soil retains too much water and becomes compacted. Causes root rot in pigeon pea.",
     "Add sand to improve drainage. Use raised beds. Consider alternative crops.",
     "Add sand and organic matter. Apply gypsum.",
     "Avoid waterlogging. Ensure good drainage."],
    ["Pigeon Pea", "Loamy", "HIGH",
     "Loamy soil is ideal for pigeon pea with good drainage and fertility.",
     "No major amendments needed. Apply phosphorus. Use balanced fertilizers.",
     "Maintain organic matter. Apply phosphorus.",
     "Use rainfed or light irrigation. Drought-tolerant crop."],
]
add_sheet_data(ws_rules, rules_headers, rules_data)

# Sheet 2: Validation History
ws_history = wb.create_sheet("Validation History")
history_headers = [
    "Timestamp", "User ID", "Farm Name", "Crop", "Soil Type",
    "Suitability", "Explanation", "Recommended Action", "Amendments", "Irrigation"
]
history_data = []
add_sheet_data(ws_history, history_headers, history_data)

# Sheet 3: Quick Actions Reference
ws_quick = wb.create_sheet("Quick Actions")
quick_headers = ["Suitability", "Action Type", "Message", "Priority"]
quick_data = [
    ["HIGH", "Confirmation", "This crop-soil combination is highly suitable. Proceed with standard farming practices.", "LOW"],
    ["MODERATE", "Advisory", "This combination is moderately suitable. Follow the recommended amendments for better yield.", "MEDIUM"],
    ["LOW", "Warning", "This crop is NOT suitable for the selected soil type. Consider changing crop or soil amendments.", "HIGH"],
    ["LOW", "Alternative", "Consider alternative crops better suited for this soil type.", "HIGH"],
    ["LOW", "Amendment", "Significant soil amendments required before planting. Consult an agronomist.", "HIGH"],
]
add_sheet_data(ws_quick, quick_headers, quick_data)

wb.save(EXCEL_PATH)
print(f"Farm validation Excel created at: {EXCEL_PATH}")
print(f"Sheets: {wb.sheetnames}")
