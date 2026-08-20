from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from typing import List, Optional
import openpyxl
from pathlib import Path

from app.database.session import get_db

router = APIRouter()

_EXCEL_PATH = Path(__file__).resolve().parent.parent.parent.parent / "recommendations.xlsx"


def _load_sheet(sheet_name: str) -> List[dict]:
    if not _EXCEL_PATH.exists():
        return []
    wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(item)
    wb.close()
    return rows


@router.get("/plants")
def get_plants():
    return _load_sheet("Plants")


@router.get("/diseases")
def get_diseases(plant_name: Optional[str] = None):
    rows = _load_sheet("Diseases")
    if plant_name:
        rows = [r for r in rows if r.get("Plant Name", "").lower() == plant_name.lower()]
    return rows


@router.get("/crop-soil")
def get_crop_soil(crop: Optional[str] = None, soil: Optional[str] = None):
    rows = _load_sheet("Crop Soil")
    if crop:
        rows = [r for r in rows if r.get("Crop", "").lower() == crop.lower()]
    if soil:
        rows = [r for r in rows if r.get("Soil Type", "").lower() == soil.lower()]
    return rows


@router.get("/weather-actions")
def get_weather_actions(crop: Optional[str] = None):
    rows = _load_sheet("Weather Actions")
    if crop:
        rows = [r for r in rows if r.get("Crop", "").lower() == crop.lower()]
    return rows


@router.get("/city-crops")
def get_city_crops(city: Optional[str] = None, crop: Optional[str] = None):
    rows = _load_sheet("City Crops")
    if city:
        rows = [r for r in rows if r.get("City", "").lower() == city.lower()]
    if crop:
        rows = [r for r in rows if r.get("Dominant Crop", "").lower() == crop.lower()]
    return rows


@router.get("/faq")
def get_faq(intent: Optional[str] = None):
    rows = _load_sheet("FAQ")
    if intent:
        rows = [r for r in rows if r.get("Intent", "").lower() == intent.lower()]
    return rows


@router.get("/general-agriculture")
def get_general_agriculture(topic: Optional[str] = None):
    rows = _load_sheet("General Agriculture")
    if topic:
        rows = [r for r in rows if r.get("Topic", "").lower() == topic.lower()]
    return rows
