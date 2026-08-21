import os
import uuid
import httpx
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
import openpyxl

from app.models.leaf_analysis import LeafAnalysis
from app.models.farm import Farm
from app.models.crop import Crop
from app.models.weather import WeatherObservation
from app.models.satellite import SatelliteObservation
from app.models.subscription import Subscription
from app.models.user import User
from app.ml.leaf_vision import LeafVisionService
from app.ml.model_registry import get_model_registry
from app.services.recommendation import RecommendationEngine
from app.services.disease_risk import DiseaseRiskEngine
from app.services.ai_explanation import AIExplanationService
from app.services.excel_recommendation import ExcelRecommendationService
from app.services.leaf_doctor_training import LeafDoctorTrainingService
from app.providers.ai import MockAIProvider
from app.core.config import settings


class LeafAnalysisService:
    def __init__(self) -> None:
        self.vision = LeafVisionService()
        self.registry = get_model_registry()
        self.excel_path = str(Path(__file__).resolve().parent.parent.parent.parent / "leaf_doctor_training.xlsx")

    def _record_analysis(
        self,
        user_id: int,
        farm_id: Optional[int],
        farm_name: Optional[str],
        crop: Optional[str],
        condition: Optional[str],
        severity: Optional[str],
        health_status: Optional[str],
        risk_score: Optional[float],
        risk_level: Optional[str],
        recommendation: Optional[str],
        recommendation_source: Optional[str],
        latitude: Optional[float],
        longitude: Optional[float],
        temperature: Optional[float],
        humidity: Optional[float],
        rainfall: Optional[float],
        ndvi: Optional[float],
        model_version: Optional[str],
    ) -> None:
        if not Path(self.excel_path).exists():
            return
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            if "Leaf Analysis History" not in wb.sheetnames:
                ws = wb.create_sheet("Leaf Analysis History")
                ws.append([
                    "Timestamp", "User ID", "Farm ID", "Farm Name", "Crop", "Condition",
                    "Severity", "Health Status", "Risk Score", "Risk Level",
                    "Recommendation", "Recommendation Source", "Latitude", "Longitude",
                    "Temperature", "Humidity", "Rainfall", "NDVI", "Model Version"
                ])
            ws = wb["Leaf Analysis History"]
            ws.append([
                datetime.now(timezone.utc).isoformat(),
                user_id,
                farm_id,
                farm_name,
                crop,
                condition.replace("_", " ") if condition else None,
                severity,
                health_status,
                risk_score,
                risk_level,
                recommendation,
                recommendation_source,
                latitude,
                longitude,
                temperature,
                humidity,
                rainfall,
                ndvi,
                model_version,
            ])
            wb.save(self.excel_path)
            wb.close()
        except Exception as exc:
            print(f"Failed to record leaf analysis to Excel: {exc}")

    def enforce_subscription_limit(self, user: User, db: Session) -> Optional[Dict[str, str]]:
        sub = db.query(Subscription).filter(Subscription.user_id == user.id).first()
        plan = (sub.plan if sub else "FREE").upper()
        limits = {
            "FREE": 3,
            "FARMER": 20,
            "PROFESSIONAL": 100,
            "ENTERPRISE": 999999,
        }
        limit = limits.get(plan, 3)
        used = sub.used_this_month if sub else 0
        if used >= limit:
            return {"detail": f"Monthly leaf analysis limit reached ({limit}). Please upgrade your plan."}
        return None

    def increment_usage(self, user: User, db: Session) -> None:
        sub = db.query(Subscription).filter(Subscription.user_id == user.id).first()
        if sub:
            sub.used_this_month = (sub.used_this_month or 0) + 1
            db.add(sub)

    def _save_image(self, file_bytes: bytes, filename: str) -> str:
        upload_dir = os.path.join("uploads", "leaf")
        os.makedirs(upload_dir, exist_ok=True)
        ext = os.path.splitext(filename)[1] or ".jpg"
        name = f"{uuid.uuid4().hex}{ext}"
        path = os.path.join(upload_dir, name)
        with open(path, "wb") as f:
            f.write(file_bytes)
        return f"/static/leaf/{name}"

    def _get_farm_context(self, farm_id: Optional[int], db: Session) -> Dict[str, Any]:
        if not farm_id:
            return {}
        farm = db.query(Farm).filter(Farm.id == farm_id).first()
        if not farm:
            return {}
        crop = db.query(Crop).filter(Crop.id == farm.crop_id).first()
        return {
            "farm_id": farm.id,
            "farm_name": farm.name,
            "crop": crop.name if crop else None,
            "crop_scientific_name": crop.scientific_name if crop else None,
            "growth_stage": None,
            "latitude": farm.latitude,
            "longitude": farm.longitude,
            "sowing_date": farm.sowing_date.isoformat() if farm.sowing_date else None,
        }

    def _get_weather_context(self, farm_id: Optional[int], db: Session) -> Dict[str, Any]:
        if not farm_id:
            return {}
        weather = (
            db.query(WeatherObservation)
            .filter(WeatherObservation.farm_id == farm_id)
            .order_by(WeatherObservation.recorded_at.desc())
            .first()
        )
        if not weather:
            return {}
        return {
            "temperature": weather.temperature,
            "humidity": weather.humidity,
            "rainfall": weather.rainfall,
            "wind_speed": weather.wind_speed,
        }

    def _get_satellite_context(self, farm_id: Optional[int], db: Session) -> Dict[str, Any]:
        if not farm_id:
            return {}
        latest = (
            db.query(SatelliteObservation)
            .filter(SatelliteObservation.farm_id == farm_id)
            .order_by(SatelliteObservation.observation_date.desc())
            .first()
        )
        if not latest:
            return {}
        history = (
            db.query(SatelliteObservation)
            .filter(SatelliteObservation.farm_id == farm_id)
            .order_by(SatelliteObservation.observation_date.desc())
            .limit(5)
            .all()
        )
        ndvi_trend = 0.0
        if len(history) >= 2:
            ndvi_trend = (history[0].ndvi or 0) - (history[1].ndvi or 0)
        return {
            "ndvi": latest.ndvi,
            "ndmi": latest.ndmi,
            "ndwi": latest.ndwi,
            "cloud_percentage": latest.cloud_percentage,
            "observation_date": latest.observation_date.isoformat() if latest.observation_date else None,
            "ndvi_trend": ndvi_trend,
        }

    def _get_live_weather(self, latitude: float, longitude: float) -> Dict[str, Any]:
        try:
            with httpx.Client() as client:
                resp = client.get(
                    "https://api.open-meteo.com/v1/forecast",
                    params={
                        "latitude": latitude,
                        "longitude": longitude,
                        "current": "temperature_2m,relative_humidity_2m,precipitation,wind_speed_10m",
                        "timezone": "auto",
                    },
                    timeout=30,
                )
                if resp.status_code == 200:
                    data = resp.json()
                    current = data.get("current", {})
                    return {
                        "temperature": current.get("temperature_2m"),
                        "humidity": current.get("relative_humidity_2m"),
                        "rainfall": current.get("precipitation"),
                        "wind_speed": current.get("wind_speed_10m"),
                        "source": "open_meteo_live",
                    }
        except Exception:
            pass
        return {}

    def _calculate_severity(
        self,
        disease_confidence: float,
        risk_level: str,
        ndvi_trend: float,
        condition: str,
    ) -> str:
        if condition and condition.lower() == "healthy":
            return "LOW"
        score = 0.0
        if disease_confidence >= 0.9:
            score += 30
        elif disease_confidence >= 0.75:
            score += 20
        elif disease_confidence >= 0.5:
            score += 10
        if risk_level == "HIGH":
            score += 30
        elif risk_level == "MODERATE":
            score += 20
        if ndvi_trend < -0.05:
            score += 20
        elif ndvi_trend < -0.02:
            score += 10
        if score >= 60:
            return "HIGH"
        if score >= 35:
            return "MODERATE"
        return "LOW"

    def _calculate_health_status(self, severity: str, risk_level: str, condition: str) -> str:
        if condition and condition.lower() == "healthy":
            return "Healthy"
        if severity == "HIGH" or risk_level == "HIGH":
            return "At Risk"
        if severity == "MODERATE" or risk_level == "MODERATE":
            return "Needs Attention"
        return "Monitor"

    def _determine_risk_from_leaf(
        self,
        condition: Optional[str],
        disease_confidence: float,
        severity: str,
        weather: Dict[str, Any],
        satellite: Dict[str, Any],
    ) -> Dict[str, Any]:
        score = 0.0
        factors: List[str] = []

        if condition and condition.lower() != "healthy":
            score += 20
            factors.append(f"leaf condition: {condition.replace('_', ' ')}")

        if disease_confidence >= 0.8:
            score += 20
            factors.append("high model confidence")
        elif disease_confidence >= 0.5:
            score += 10
            factors.append("moderate model confidence")

        if severity == "HIGH":
            score += 25
            factors.append("high severity indicators")
        elif severity == "MODERATE":
            score += 15
            factors.append("moderate severity indicators")

        humidity = weather.get("humidity")
        if humidity is not None and humidity > 80:
            score += 15
            factors.append("high humidity")
        elif humidity is not None and humidity > 65:
            score += 5

        temperature = weather.get("temperature")
        if temperature is not None and 20 <= temperature <= 30:
            score += 10
            factors.append("favorable temperature for disease")

        ndvi = satellite.get("ndvi")
        if ndvi is not None and ndvi < 0.3:
            score += 15
            factors.append("low vegetation index")

        if score >= 50:
            level = "HIGH"
        elif score >= 25:
            level = "MODERATE"
        else:
            level = "LOW"

        explanation = (
            "Leaf analysis and environmental conditions suggest "
            + ", ".join(factors)
            + ". Inspect leaves and consult an agronomist before applying treatment."
            if factors
            else "Current conditions do not strongly indicate elevated disease risk."
        )

        return {
            "risk_type": "LEAF_DISEASE_RISK",
            "risk_level": level,
            "score": round(score, 1),
            "explanation": explanation,
        }

    def _build_recommendation(
        self, risk: Dict[str, Any], condition: Optional[str], crop: Optional[str]
    ) -> Dict[str, Any]:
        engine = RecommendationEngine()
        recs = engine.generate([risk], {"crop": crop or "Unknown", "condition": condition or "Unknown"})
        return recs[0] if recs else {
            "priority": "MEDIUM",
            "title": "General Advisory",
            "recommendation": "Monitor crop condition and consult an agronomist if symptoms worsen.",
            "reasoning": risk.get("explanation", ""),
        }

    def _identify_crop_with_gemini(self, file_bytes: bytes, filename: str) -> Optional[str]:
        try:
            provider = LLMProvider() if (settings.AI_PROVIDER == "gemini" and settings.GEMINI_API_KEY) else MockAIProvider()
            if not hasattr(provider, "generate_image_explanation"):
                return None
            ext = Path(filename).suffix.lower()
            mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', '.webp': 'image/webp'}
            mime_type = mime_map.get(ext, 'image/jpeg')
            result = provider.generate_image_explanation(
                file_bytes,
                mime_type=mime_type,
                prompt="Identify the plant/crop from this leaf image. Return ONLY the exact plant name. Do not make up names. If unsure, return 'Unknown'."
            )
            explanation = result.get("explanation", "") or ""
            if "does not support image input" in explanation.lower() or "cannot read" in explanation.lower():
                return None
            crop = explanation.split("\n")[0].strip().split(",")[0].strip()
            crop = crop.replace("_", " ").strip()
            if crop and len(crop) < 50 and crop.lower() != "unknown":
                return crop
        except Exception:
            pass
        return None

    def analyze_leaf(
        self,
        user: User,
        file_bytes: bytes,
        filename: str,
        farm_id: Optional[int],
        crop_override: Optional[str],
        db: Session,
        latitude: Optional[float] = None,
        longitude: Optional[float] = None,
    ) -> Dict[str, Any]:
        limit_error = self.enforce_subscription_limit(user, db)
        if limit_error:
            return limit_error

        valid, err = self.vision.validate_image(file_bytes, filename)
        if not valid:
            return {"status": "error", "message": err}

        image_url = self._save_image(file_bytes, filename)
        try:
            analysis_result = self.vision.analyze(file_bytes, crop_override=crop_override, filename=filename)
            if not analysis_result or analysis_result.get("status") == "error":
                return analysis_result or {"status": "error", "message": "Analysis failed"}
        except Exception as exc:
            record = LeafAnalysis(
                user_id=user.id,
                farm_id=farm_id,
                image_url=image_url,
                status="error",
                error_message=str(exc),
                created_at=datetime.now(timezone.utc),
            )
            db.add(record)
            db.commit()
            return {"status": "error", "message": str(exc)}

        if analysis_result.get("status") == "uncertain":
            gemini_crop = self._identify_crop_with_gemini(file_bytes, filename)
            if gemini_crop:
                analysis_result = self.vision.analyze(file_bytes, crop_override=gemini_crop, filename=filename)
            else:
                analysis_result["status"] = "success"
                analysis_result["crop"] = analysis_result.get("crop") or "Unknown"
                analysis_result["crop_confidence"] = 0.0
                analysis_result["condition"] = analysis_result.get("condition") or "healthy"
                analysis_result["disease_confidence"] = 0.0
                analysis_result["model_version"] = analysis_result.get("model_version") or "fallback"
                analysis_result["low_confidence"] = True

        if analysis_result.get("status") != "success":
            record = LeafAnalysis(
                user_id=user.id,
                farm_id=farm_id,
                image_url=image_url,
                status=analysis_result.get("status", "error"),
                error_message=analysis_result.get("message"),
                created_at=datetime.now(timezone.utc),
            )
            db.add(record)
            db.commit()
            return analysis_result

        crop = analysis_result.get("crop")
        crop_confidence = analysis_result.get("crop_confidence")
        condition = analysis_result.get("condition")
        disease_confidence = analysis_result.get("disease_confidence")
        model_version = analysis_result.get("model_version")
        crop_model_version = analysis_result.get("crop_model_version")
        low_confidence = analysis_result.get("low_confidence", False)

        farm_context = self._get_farm_context(farm_id, db)
        weather = self._get_weather_context(farm_id, db)

        if latitude is not None and longitude is not None:
            live_weather = self._get_live_weather(latitude, longitude)
            if live_weather:
                weather = live_weather

        satellite = self._get_satellite_context(farm_id, db)

        risk = self._determine_risk_from_leaf(
            condition=condition,
            disease_confidence=disease_confidence or 0.0,
            severity="",
            weather=weather,
            satellite=satellite,
        )

        severity = self._calculate_severity(
            disease_confidence=disease_confidence or 0.0,
            risk_level=risk["risk_level"],
            ndvi_trend=satellite.get("ndvi_trend", 0.0),
            condition=condition or "",
        )
        risk["severity"] = severity

        health_status = self._calculate_health_status(severity, risk["risk_level"], condition or "")
        recommendation = self._build_recommendation(risk, condition, crop)

        training_text: Optional[str] = None
        recommendation_source = "knowledge_base"

        try:
            training_service = LeafDoctorTrainingService()
            training_match = training_service.get_training_match(crop, condition)
            if training_match:
                training_text = training_service.get_recommendation(crop, condition)
                recommendation_source = "training_data"
            else:
                ai_service = AIExplanationService()
                if hasattr(ai_service.provider, "generate_image_explanation"):
                    ext = Path(filename).suffix.lower()
                    mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', '.webp': 'image/webp'}
                    mime_type = mime_map.get(ext, 'image/jpeg')
                    gemini_result = ai_service.provider.generate_image_explanation(file_bytes, mime_type=mime_type)
                    if gemini_result and gemini_result.get("explanation"):
                        training_text = gemini_result.get("explanation")
                        recommendation_source = "gemini"
        except Exception:
            pass

        final_recommendation = training_text or recommendation.get("recommendation", "")

        record = LeafAnalysis(
            user_id=user.id,
            farm_id=farm_id,
            image_url=image_url,
            crop=crop,
            crop_confidence=crop_confidence,
            condition=condition,
            disease_confidence=disease_confidence,
            severity=severity,
            health_status=health_status,
            symptoms=condition.replace("_", " ") if condition else None,
            ndvi=satellite.get("ndvi"),
            ndmi=satellite.get("ndmi"),
            ndwi=satellite.get("ndwi"),
            temperature=weather.get("temperature"),
            humidity=weather.get("humidity"),
            rainfall=weather.get("rainfall"),
            risk_score=risk["score"],
            risk_level=risk["risk_level"],
            model_version=model_version,
            crop_model_version=crop_model_version,
            recommendation=final_recommendation,
            recommendation_source=recommendation_source,
            status="success",
            latitude=latitude,
            longitude=longitude,
            created_at=datetime.now(timezone.utc),
        )
        db.add(record)
        self.increment_usage(user, db)

        db.commit()
        db.refresh(record)

        self._record_analysis(
            user_id=user.id,
            farm_id=farm_id,
            farm_name=farm_context.get("farm_name"),
            crop=record.crop,
            condition=record.condition,
            severity=record.severity,
            health_status=record.health_status,
            risk_score=record.risk_score,
            risk_level=record.risk_level,
            recommendation=record.recommendation,
            recommendation_source=record.recommendation_source,
            latitude=latitude,
            longitude=longitude,
            temperature=record.temperature,
            humidity=record.humidity,
            rainfall=record.rainfall,
            ndvi=record.ndvi,
            model_version=record.model_version,
        )

        return {
            "status": "success",
            "id": record.id,
            "crop": record.crop,
            "crop_confidence": record.crop_confidence,
            "condition": record.condition,
            "disease_confidence": record.disease_confidence,
            "severity": record.severity,
            "health_status": record.health_status,
            "symptoms": record.symptoms,
            "ndvi": record.ndvi,
            "ndmi": record.ndmi,
            "ndwi": record.ndwi,
            "temperature": record.temperature,
            "humidity": record.humidity,
            "rainfall": record.rainfall,
            "risk_score": record.risk_score,
            "risk_level": record.risk_level,
            "model_version": record.model_version,
            "crop_model_version": record.crop_model_version,
            "recommendation": record.recommendation,
            "recommendation_source": record.recommendation_source,
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "latitude": latitude,
            "longitude": longitude,
            "weather_source": weather.get("source") if isinstance(weather, dict) else None,
            "low_confidence": low_confidence,
        }
