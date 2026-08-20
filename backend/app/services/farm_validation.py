from typing import Dict, Any, Optional, List
from pathlib import Path
import openpyxl
from datetime import datetime, timezone


class FarmValidationService:
    def __init__(self, excel_path: Optional[str] = None) -> None:
        if excel_path is None:
            excel_path = str(Path(__file__).resolve().parent.parent.parent.parent / "farm_validation.xlsx")
        self.excel_path = excel_path
        self._cache: Optional[Dict[str, Dict[str, Any]]] = None

    def _load(self) -> Dict[str, Dict[str, Any]]:
        if self._cache is not None:
            return self._cache
        if not Path(self.excel_path).exists():
            self._cache = {}
            return self._cache
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        data: Dict[str, Dict[str, Any]] = {}
        if "Crop Soil Rules" in wb.sheetnames:
            ws = wb["Crop Soil Rules"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {headers[i]: row[i] for i in range(len(headers))}
                crop = str(item.get("Crop") or "").strip().lower()
                soil = str(item.get("Soil Type") or "").strip().lower()
                key = f"{crop}|{soil}"
                data[key] = item
        wb.close()
        self._cache = data
        return data

    def validate(self, crop: Optional[str], soil_type: Optional[str]) -> Dict[str, Any]:
        if not crop or not soil_type:
            return {
                "crop": crop,
                "soil_type": soil_type,
                "suitability": "UNKNOWN",
                "explanation": "Crop or soil type missing.",
                "recommended_action": "Please select both crop and soil type.",
                "amendments_required": "",
                "irrigation_adjustment": "",
            }
        data = self._load()
        crop_key = str(crop).strip().lower()
        soil_key = str(soil_type).strip().lower()
        exact = data.get(f"{crop_key}|{soil_key}")
        if exact:
            return {
                "crop": crop,
                "soil_type": soil_type,
                "suitability": str(exact.get("Suitability") or "UNKNOWN"),
                "explanation": str(exact.get("Explanation") or ""),
                "recommended_action": str(exact.get("Recommended Action") or ""),
                "amendments_required": str(exact.get("Amendments Required") or ""),
                "irrigation_adjustment": str(exact.get("Irrigation Adjustment") or ""),
            }
        for key, item in data.items():
            item_crop = str(item.get("Crop") or "").strip().lower()
            item_soil = str(item.get("Soil Type") or "").strip().lower()
            if item_crop == crop_key and soil_key in item_soil:
                return {
                    "crop": crop,
                    "soil_type": soil_type,
                    "suitability": str(item.get("Suitability") or "UNKNOWN"),
                    "explanation": str(item.get("Explanation") or ""),
                    "recommended_action": str(item.get("Recommended Action") or ""),
                    "amendments_required": str(item.get("Amendments Required") or ""),
                    "irrigation_adjustment": str(item.get("Irrigation Adjustment") or ""),
                }
        return {
            "crop": crop,
            "soil_type": soil_type,
            "suitability": "UNKNOWN",
            "explanation": f"No validation data found for {crop} in {soil_type}.",
            "recommended_action": "Consult local agricultural extension for guidance.",
            "amendments_required": "",
            "irrigation_adjustment": "",
        }

    def record_validation(self, user_id: int, farm_name: str, crop: str, soil_type: str, result: Dict[str, Any]) -> None:
        if not Path(self.excel_path).exists():
            return
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            if "Validation History" not in wb.sheetnames:
                ws = wb.create_sheet("Validation History")
                ws.append([
                    "Timestamp", "User ID", "Farm Name", "Crop", "Soil Type",
                    "Suitability", "Explanation", "Recommended Action", "Amendments", "Irrigation"
                ])
            ws = wb["Validation History"]
            ws.append([
                datetime.now(timezone.utc).isoformat(),
                user_id,
                farm_name,
                crop,
                soil_type,
                result.get("suitability"),
                result.get("explanation"),
                result.get("recommended_action"),
                result.get("amendments_required"),
                result.get("irrigation_adjustment"),
            ])
            wb.save(self.excel_path)
            wb.close()
        except Exception as exc:
            print(f"Failed to record validation to Excel: {exc}")
