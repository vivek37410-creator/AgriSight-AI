import re
import math
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path

import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from app.core.config import settings


_KNOWLEDGE_PATH = Path(__file__).resolve().parent.parent / "knowledge" / "agrisight_knowledge.xlsx"

_STOP_WORDS = {
    "the", "and", "or", "a", "an", "is", "are", "was", "were", "be", "been", "being",
    "have", "has", "had", "do", "does", "did", "will", "would", "could", "should", "may",
    "might", "shall", "can", "to", "of", "in", "for", "on", "with", "as", "by", "at",
    "from", "into", "about", "through", "during", "before", "after", "above", "below",
    "between", "under", "again", "further", "then", "once", "here", "there", "when",
    "where", "why", "how", "all", "each", "every", "both", "few", "more", "most", "other",
    "some", "such", "no", "nor", "not", "only", "own", "same", "so", "than", "too", "very",
    "just", "because", "but", "if", "while", "that", "this", "these", "those", "it", "its",
    "my", "your", "his", "her", "their", "our", "what", "which", "who", "whom", "am",
    "me", "he", "she", "they", "them", "us", "i", "we", "you", "get", "got", "make",
    "made", "take", "took", "see", "saw", "know", "knew", "think", "thought", "say",
    "said", "tell", "told", "ask", "asked", "want", "wanted", "use", "used", "like",
    "liked", "help", "helped", "try", "tried", "call", "called", "try", "tried", "need",
    "needed", "feel", "felt", "become", "became", "leave", "left", "put", "put", "mean",
    "meant", "keep", "kept", "let", "let", "begin", "began", "seem", "seemed", "help",
    "show", "showed", "hear", "heard", "play", "played", "run", "ran", "move", "moved",
    "live", "lived", "believe", "believed", "bring", "brought", "happen", "happened",
    "write", "wrote", "provide", "provided", "sit", "sat", "stand", "stood", "lose",
    "lost", "pay", "paid", "meet", "met", "include", "included", "continue", "continued",
    "set", "set", "learn", "learned", "change", "changed", "lead", "led", "understand",
    "understood", "watch", "watched", "follow", "followed", "stop", "stopped", "create",
    "created", "speak", "spoke", "read", "read", "allow", "allowed", "add", "added",
    "spend", "spent", "grow", "grew", "open", "opened", "walk", "walked", "win", "won",
    "offer", "offered", "remember", "remembered", "consider", "considered", "appear",
    "appeared", "buy", "bought", "wait", "waited", "serve", "served", "die", "died",
    "send", "sent", "expect", "expected", "build", "built", "stay", "stayed", "fall",
    "fell", "cut", "cut", "reach", "reached", "kill", "killed", "remain", "remained",
    "suggest", "suggested", "raise", "raised", "pass", "passed", "sell", "sold", "require",
    "required", "report", "reported", "decide", "decided", "pull", "pulled", "develop",
    "developed", "eat", "ate", "break", "broke", "blow", "blew", "close", "closed",
    "drive", "drove", "drink", "drank", "feed", "fed", "fight", "fought", "find", "found",
    "fly", "flew", "forget", "forgot", "forgive", "forgave", "freeze", "froze", "hang",
    "hung", "hide", "hid", "hit", "hit", "hurt", "hurt", "lay", "laid", "lie", "lay",
    "light", "lit", "ring", "rang", "rise", "rose", "shake", "shook", "shine", "shone",
    "shoot", "shot", "sing", "sang", "sink", "sank", "sleep", "slept", "slide", "slid",
    "speak", "spoke", "steal", "stole", "swim", "swam", "take", "took", "teach", "taught",
    "tear", "tore", "throw", "threw", "wear", "wore", "weep", "wept", "will", "would",
    "shall", "should", "may", "might", "can", "could", "must", "ought", "need", "dare",
    "used", "going", "coming", "looking", "trying", "giving", "getting", "making", "taking",
    "seeing", "knowing", "thinking", "saying", "telling", "asking", "wanting", "using",
    "liking", "helping", "trying", "calling", "needing", "feeling", "becoming", "leaving",
    "putting", "meaning", "keeping", "letting", "beginning", "showing", "hearing", "playing",
    "running", "moving", "living", "believing", "bringing", "happening", "writing",
    "providing", "sitting", "standing", "losing", "paying", "meeting", "including",
    "continuing", "setting", "learning", "changing", "leading", "understanding", "watching",
    "following", "stopping", "creating", "speaking", "reading", "allowing", "adding",
    "spending", "growing", "opening", "walking", "winning", "offering", "remembering",
    "considering", "appearing", "buying", "waiting", "serving", "dying", "sending",
    "expecting", "building", "staying", "falling", "cutting", "reaching", "killing",
    "remaining", "suggesting", "raising", "passing", "selling", "requiring", "reporting",
    "deciding", "pulling", "developing", "eating", "breaking", "blowing", "closing",
    "driving", "drinking", "feeding", "fighting", "finding", "flying", "forgetting",
    "forgiving", "freezing", "hanging", "hiding", "hitting", "hurting", "laying", "lying",
    "lighting", "ringing", "rising", "shaking", "shining", "shooting", "singing", "sinking",
    "sleeping", "sliding", "stealing", "swimming", "taking", "teaching", "tearing",
    "throwing", "wearing", "weeping", "my", "your", "his", "her", "its", "our", "their",
    "mine", "yours", "hers", "ours", "theirs", "myself", "yourself", "himself", "herself",
    "itself", "ourselves", "themselves", "who", "whom", "whose", "which", "that", "this",
    "these", "those", "what", "whatever", "whoever", "whomever", "whosever", "whichever",
    "whatever", "whomever", "whosever", "whichever", "about", "above", "across", "after",
    "against", "along", "among", "around", "at", "before", "behind", "below", "beneath",
    "beside", "between", "beyond", "by", "despite", "down", "during", "except", "for",
    "from", "in", "inside", "into", "like", "near", "of", "off", "on", "onto", "out",
    "outside", "over", "past", "since", "through", "throughout", "till", "toward", "under",
    "underneath", "until", "up", "upon", "with", "within", "without",
}


class KnowledgeEntry:
    def __init__(self, row: List[Any]):
        self.id = int(row[0]) if row[0] else 0
        self.category = str(row[1] or "").strip()
        self.crop = str(row[2] or "").strip()
        self.topic = str(row[3] or "").strip()
        self.question = str(row[4] or "").strip()
        self.keywords = str(row[5] or "").strip()
        self.answer = str(row[6] or "").strip()
        self.recommendation = str(row[7] or "").strip()
        self.severity = str(row[8] or "").strip()
        self.language = str(row[9] or "English").strip()

    def to_search_text(self) -> str:
        parts = [
            self.question,
            self.keywords,
            self.topic,
            self.crop,
            self.category,
            self.answer,
            self.recommendation,
        ]
        return " ".join(parts)


class ExcelAssistantService:
    def __init__(self, knowledge_path: Optional[str] = None):
        self.knowledge_path = Path(knowledge_path) if knowledge_path else _KNOWLEDGE_PATH
        self._entries: List[KnowledgeEntry] = []
        self._vectorizer: Optional[TfidfVectorizer] = None
        self._tfidf_matrix = None
        self._load_knowledge()

    def _preprocess(self, text: str) -> str:
        text = text.lower()
        text = re.sub(r"[^a-z0-9\s]", " ", text)
        tokens = text.split()
        tokens = [t for t in tokens if t not in _STOP_WORDS and len(t) > 2]
        return " ".join(tokens)

    def _load_knowledge(self):
        if not self.knowledge_path.exists():
            raise FileNotFoundError(f"Knowledge base not found at {self.knowledge_path}")
        wb = openpyxl.load_workbook(self.knowledge_path, read_only=True, data_only=True)
        ws = wb.active if "FAQ" not in wb.sheetnames else wb["FAQ"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        self._entries = [KnowledgeEntry(r) for r in rows if any(r)]
        corpus = [self._preprocess(e.to_search_text()) for e in self._entries]
        self._vectorizer = TfidfVectorizer(ngram_range=(1, 2), max_features=5000)
        self._tfidf_matrix = self._vectorizer.fit_transform(corpus)

    def reload(self):
        self._load_knowledge()

    def ask(self, question: str, crop: Optional[str] = None, language: str = "en") -> Dict[str, Any]:
        if not question.strip():
            return {
                "success": False,
                "answer": "Please enter a question.",
                "confidence": 0.0,
            }

        processed_query = self._preprocess(question)
        query_vec = self._vectorizer.transform([processed_query])
        similarities = cosine_similarity(query_vec, self._tfidf_matrix)[0]

        scored: List[Tuple[float, int]] = []
        for idx, entry in enumerate(self._entries):
            score = float(similarities[idx])
            if crop:
                crop_lower = crop.lower()
                entry_crop = entry.crop.lower()
                if crop_lower in entry_crop or entry_crop in crop_lower:
                    score = score * 1.25
                if crop_lower in self._preprocess(entry.keywords):
                    score = score * 1.15
            scored.append((min(score, 1.0), idx))
        scored.sort(key=lambda x: x[0], reverse=True)

        top_score = scored[0][0] if scored else 0.0
        threshold = getattr(settings, "ASSISTANT_CONFIDENCE_THRESHOLD", 0.35)

        if top_score < threshold:
            return {
                "success": False,
                "answer": "I couldn't find a reliable answer in the AgriSight knowledge base.\n\nPlease try asking about:\n• Crop diseases\n• Irrigation\n• Soil\n• Fertilizers\n• Pests\n• Weather\n• Crop growth",
                "confidence": round(top_score, 2),
                "category": None,
                "topic": None,
                "crop": crop,
                "recommendation": None,
                "severity": None,
            }

        best_idx = scored[0][1]
        best = self._entries[best_idx]

        alt_matches = []
        for score, idx in scored[1:4]:
            if score >= threshold * 0.7:
                alt_matches.append({
                    "id": self._entries[idx].id,
                    "crop": self._entries[idx].crop,
                    "topic": self._entries[idx].topic,
                    "confidence": round(score, 2),
                })

        return {
            "success": True,
            "answer": best.answer,
            "recommendation": best.recommendation,
            "category": best.category,
            "topic": best.topic,
            "crop": best.crop,
            "confidence": round(top_score, 2),
            "severity": best.severity,
            "alternatives": alt_matches if alt_matches else None,
        }


_assistant_service: Optional[ExcelAssistantService] = None


def get_assistant_service() -> ExcelAssistantService:
    global _assistant_service
    if _assistant_service is None:
        _assistant_service = ExcelAssistantService()
    return _assistant_service
