from typing import Dict, Any, Optional
from pathlib import Path
import openpyxl


class LeafDoctorTrainingService:
    def __init__(self, excel_path: Optional[str] = None) -> None:
        if excel_path is None:
            excel_path = str(Path(__file__).resolve().parent.parent.parent.parent / "leaf_doctor_training.xlsx")
        self.excel_path = excel_path
        self._cache: Optional[Dict[str, Dict[str, Any]]] = None

    def _load(self) -> Dict[str, Dict[str, Any]]:
        if self._cache is not None:
            return self._cache
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        data: Dict[str, Dict[str, Any]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = {headers[i]: row[i] for i in range(len(headers))}
                crop = str(item.get("Crop") or "").strip().lower()
                condition = str(item.get("Condition") or "").strip().lower()
                key = f"{crop}|{condition}"
                data[key] = item
                if crop:
                    data.setdefault(f"crop::{crop}", {})
                    data[f"crop::{crop}"].update({k: v for k, v in item.items() if k not in ("Crop", "Condition")})
        wb.close()
        self._cache = data
        return data

    def get_training_match(self, crop: Optional[str], condition: Optional[str]) -> Optional[Dict[str, Any]]:
        if not crop or not condition:
            return None
        data = self._load()
        crop_key = str(crop).strip().lower()
        condition_key = str(condition).strip().lower()
        exact = data.get(f"{crop_key}|{condition_key}")
        if exact:
            return exact
        for key, item in data.items():
            if key.startswith("crop::"):
                continue
            item_crop = str(item.get("Crop") or "").strip().lower()
            item_condition = str(item.get("Condition") or "").strip().lower()
            if item_crop == crop_key and condition_key in item_condition:
                return item
        return None

    def get_recommendation(self, crop: Optional[str], condition: Optional[str]) -> Optional[str]:
        match = self.get_training_match(crop, condition)
        if not match:
            return None
        lines = [
            f"Crop: {match.get('Crop')}",
            f"Condition: {match.get('Condition')}",
            f"Symptoms: {match.get('Symptoms')}",
            f"Causes: {match.get('Causes')}",
            f"Prevention: {match.get('Prevention')}",
            f"Recommended Action: {match.get('Recommended Action')}",
            f"Severity: {match.get('Severity')}",
        ]
        return "\n".join([l for l in lines if l and l != "None"])
