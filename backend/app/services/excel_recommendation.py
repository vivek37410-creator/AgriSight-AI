from typing import Dict, Any, Optional
from pathlib import Path
import openpyxl


class ExcelRecommendationService:
    def __init__(self, excel_path: Optional[str] = None) -> None:
        if excel_path is None:
            excel_path = str(Path(__file__).resolve().parent.parent.parent.parent / "recommendations.xlsx")
        self.excel_path = excel_path
        self._cache: Optional[Dict[str, Dict[str, Any]]] = None

    def _load(self) -> Dict[str, Dict[str, Any]]:
        if self._cache is not None:
            return self._cache
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data: Dict[str, Dict[str, Any]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: row[i] for i in range(len(headers))}
            plant_name = item.get("Plant Name")
            if plant_name:
                data[str(plant_name).strip().lower()] = item
        wb.close()
        self._cache = data
        return data

    def get_recommendation(self, plant_name: Optional[str], condition: Optional[str] = None) -> Optional[str]:
        if not plant_name:
            return None
        data = self._load()
        key = str(plant_name).strip().lower()
        item = data.get(key)
        if not item:
            return None
        lines = [
            f"Plant: {item.get('Plant Name')}",
            f"Type: {item.get('Type')}",
            f"Condition: {(condition or 'Unknown').replace('_', ' ').title()}",
        ]
        if item.get("Watering"):
            lines.append(f"Watering: {item.get('Watering')}")
        if item.get("Fertilization"):
            lines.append(f"Fertilization: {item.get('Fertilization')}")
        if item.get("Pest Control"):
            lines.append(f"Pest Control: {item.get('Pest Control')}")
        if item.get("Disease Prevention"):
            lines.append(f"Disease Prevention: {item.get('Disease Prevention')}")
        if item.get("Soil Requirements"):
            lines.append(f"Soil: {item.get('Soil Requirements')}")
        if item.get("Sunlight"):
            lines.append(f"Sunlight: {item.get('Sunlight')}")
        if item.get("General Care"):
            lines.append(f"Care: {item.get('General Care')}")
        return "\n".join(lines)
