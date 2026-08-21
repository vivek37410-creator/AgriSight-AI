from typing import Dict, Any, Optional, List
from pathlib import Path
import openpyxl


class CropSoilRecommendationService:
    def __init__(self, excel_path: Optional[str] = None) -> None:
        if excel_path is None:
            excel_path = str(Path(__file__).resolve().parent.parent.parent.parent / "recommendations.xlsx")
        self.excel_path = excel_path
        self._crop_soil_cache: Optional[List[Dict[str, Any]]] = None
        self._weather_cache: Optional[List[Dict[str, Any]]] = None

    def _load_crop_soil(self) -> List[Dict[str, Any]]:
        if self._crop_soil_cache is not None:
            return self._crop_soil_cache
        if not Path(self.excel_path).exists():
            return []
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        if "Crop Soil" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Crop Soil"]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            rows.append(item)
        wb.close()
        self._crop_soil_cache = rows
        return rows

    def _load_weather_actions(self) -> List[Dict[str, Any]]:
        if self._weather_cache is not None:
            return self._weather_cache
        if not Path(self.excel_path).exists():
            return []
        wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        if "Weather Actions" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Weather Actions"]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            rows.append(item)
        wb.close()
        self._weather_cache = rows
        return rows

    def get_crop_soil_recommendations(self, crop: Optional[str], soil_type: Optional[str]) -> List[Dict[str, Any]]:
        if not crop:
            return []
        crop_lower = crop.strip().lower()
        soil_lower = (soil_type or "").strip().lower()
        rows = self._load_crop_soil()
        matches = []
        for row in rows:
            row_crop = str(row.get("Crop", "")).strip().lower()
            row_soil = str(row.get("Soil Type", "")).strip().lower()
            if crop_lower in row_crop or row_crop in crop_lower:
                match_score = 1.0
                if soil_lower and soil_lower in row_soil:
                    match_score = 1.5
                matches.append({**row, "_match_score": match_score})
        matches.sort(key=lambda x: x.get("_match_score", 0), reverse=True)
        recommendations = []
        for match in matches[:3]:
            is_suitable = soil_lower and soil_lower in str(match.get("Soil Type", "")).lower()
            rec = {
                "priority": "MEDIUM",
                "title": f"Soil Compatibility: {crop} on {match.get('Soil Type', 'various soils')}",
                "recommendation": (
                    f"{crop} is {'well-suited' if is_suitable else 'potentially suitable'} for {match.get('Soil Type', 'this soil type')}. "
                    f"pH Range: {match.get('pH Range', 'N/A')}. "
                    f"Nutrient Needs: {match.get('Nutrient Needs', 'N/A')}. "
                    f"Irrigation: {match.get('Irrigation', 'N/A')}."
                ),
                "reasoning": f"Based on crop-soil compatibility data for {crop}.",
            }
            if not is_suitable and soil_lower:
                rec["priority"] = "HIGH"
                rec["title"] = f"Soil Mismatch Warning: {crop}"
                rec["recommendation"] = (
                    f"{crop} is typically grown on {match.get('Soil Type', 'different soil')}. "
                    f"Your current soil type is {soil_type}. This may affect yield. "
                    f"Consider soil amendments or switching to a crop better suited for {soil_type}."
                )
            recommendations.append(rec)
        return recommendations

    def get_weather_recommendations(self, crop: Optional[str], weather: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not crop or not weather:
            return []
        crop_lower = crop.strip().lower()
        rows = self._load_weather_actions()
        matches = []
        for row in rows:
            row_crop = str(row.get("Crop", "")).strip().lower()
            if crop_lower in row_crop or row_crop in crop_lower:
                matches.append(row)
        recommendations = []
        for match in matches[:2]:
            risk_level = str(match.get("Risk Level", "")).upper()
            priority = "MEDIUM"
            if risk_level == "HIGH":
                priority = "HIGH"
            elif risk_level == "LOW":
                priority = "LOW"
            recommendations.append({
                "priority": priority,
                "title": f"Weather Advisory: {match.get('Weather Condition', 'General')}",
                "recommendation": match.get("Recommended Action", ""),
                "reasoning": f"Matched weather action for {crop} under {match.get('Weather Condition', 'general conditions')}.",
            })
        return recommendations
